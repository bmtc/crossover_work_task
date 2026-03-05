#!/usr/bin/env python3
"""
Add Potential Savings Columns to Vendor Analysis
Calculates realistic savings by recommendation type and consolidation group
"""

import openpyxl
from collections import defaultdict
import shutil
import os
import time

file_path = "output/output_file.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb['Vendor Analysis Assessment']

print("=" * 140)
print("ADDING POTENTIAL SAVINGS COLUMNS")
print("=" * 140)

# Add headers for new columns
# Column G: Savings Category
# Column H: Estimated Annual Savings (USD)
# Column I: Savings Potential (Low/Medium/High)

ws.cell(1, 7).value = "Savings Category"
ws.cell(1, 8).value = "Estimated Annual Savings (USD)"
ws.cell(1, 9).value = "Savings Potential"

print("\nHeaders added:")
print("  Column G: Savings Category")
print("  Column H: Estimated Annual Savings (USD)")
print("  Column I: Savings Potential")

# Define savings rates by scenario
savings_rules = {
    'Terminate': {
        'rate': 1.0,  # 100% savings
        'category': 'Full Elimination',
        'potential': 'High'
    },
    'Consolidate - Duplicate': {
        'rate': 0.40,  # 40% savings (eliminate one contract)
        'category': 'Duplicate Entity Consolidation',
        'potential': 'High'
    },
    'Consolidate - Real Estate': {
        'rate': 0.15,  # 15% savings (volume consolidation)
        'category': 'Real Estate Consolidation',
        'potential': 'High'
    },
    'Consolidate - Cloud': {
        'rate': 0.12,  # 12% savings (master agreement)
        'category': 'Cloud Infrastructure Consolidation',
        'potential': 'Medium'
    },
    'Consolidate - Insurance': {
        'rate': 0.20,  # 20% savings (consolidate carriers)
        'category': 'Insurance Consolidation',
        'potential': 'Medium'
    },
    'Consolidate - Project Management': {
        'rate': 0.10,  # 10% savings (single tool)
        'category': 'Project Management Consolidation',
        'potential': 'Medium'
    },
    'Consolidate - Consulting': {
        'rate': 0.15,  # 15% savings (preferred firm)
        'category': 'Consulting Consolidation',
        'potential': 'Medium'
    },
    'Consolidate - Recruitment': {
        'rate': 0.12,  # 12% savings (preferred agencies)
        'category': 'Recruitment Consolidation',
        'potential': 'Medium'
    },
    'Consolidate - Generic': {
        'rate': 0.08,  # 8% savings (generic consolidation)
        'category': 'Vendor Consolidation',
        'potential': 'Low'
    },
    'Optimize - High Spend': {
        'rate': 0.12,  # 12% savings (negotiation)
        'category': 'Contract Negotiation',
        'potential': 'Medium'
    },
    'Optimize - Mid Spend': {
        'rate': 0.08,  # 8% savings (negotiation)
        'category': 'Contract Optimization',
        'potential': 'Low'
    },
    'Optimize - Low Spend': {
        'rate': 0.05,  # 5% savings (limited opportunity)
        'category': 'Usage Optimization',
        'potential': 'Low'
    }
}

# Collect vendors by consolidation group
consolidation_groups = {
    'Navan': {'vendors': [], 'rate': 0.40, 'category': 'Duplicate Entity Consolidation'},
    'Real Estate': {'vendors': [], 'rate': 0.15, 'category': 'Real Estate Consolidation'},
    'Cloud': {'vendors': [], 'rate': 0.12, 'category': 'Cloud Infrastructure Consolidation'},
    'Insurance': {'vendors': [], 'rate': 0.20, 'category': 'Insurance Consolidation'},
    'Project Management': {'vendors': [], 'rate': 0.10, 'category': 'Project Management Consolidation'},
}

# Calculate savings for each vendor
savings_total = 0
savings_count = 0

for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    spend = ws.cell(row, 3).value or 0
    desc = ws.cell(row, 4).value or ""
    rec = ws.cell(row, 5).value or ""
    note = ws.cell(row, 6).value or ""

    if not vendor_name:
        continue

    savings_category = ""
    estimated_savings = 0
    savings_potential = ""

    # Determine savings based on recommendation and context
    if rec == 'Terminate':
        # Full elimination savings
        estimated_savings = int(spend * savings_rules['Terminate']['rate'])
        savings_category = savings_rules['Terminate']['category']
        savings_potential = savings_rules['Terminate']['potential']

    elif rec == 'Consolidate':
        # Consolidation savings - depends on type
        if 'navan' in vendor_name.lower():
            # Navan duplicate - high savings (40%)
            estimated_savings = int(spend * savings_rules['Consolidate - Duplicate']['rate'])
            savings_category = savings_rules['Consolidate - Duplicate']['category']
            savings_potential = savings_rules['Consolidate - Duplicate']['potential']

        elif any(kw in desc.lower() for kw in ['real estate', 'office', 'property', 'workspace']):
            # Real estate consolidation
            estimated_savings = int(spend * savings_rules['Consolidate - Real Estate']['rate'])
            savings_category = savings_rules['Consolidate - Real Estate']['category']
            savings_potential = savings_rules['Consolidate - Real Estate']['potential']

        elif any(kw in desc.lower() for kw in ['cloud', 'infrastructure', 'hosting', 'aws', 'azure']):
            # Cloud consolidation
            estimated_savings = int(spend * savings_rules['Consolidate - Cloud']['rate'])
            savings_category = savings_rules['Consolidate - Cloud']['category']
            savings_potential = savings_rules['Consolidate - Cloud']['potential']

        elif any(kw in desc.lower() for kw in ['insurance', 'coverage', 'benefit']):
            # Insurance consolidation
            estimated_savings = int(spend * savings_rules['Consolidate - Insurance']['rate'])
            savings_category = savings_rules['Consolidate - Insurance']['category']
            savings_potential = savings_rules['Consolidate - Insurance']['potential']

        elif any(kw in desc.lower() for kw in ['project', 'task', 'tracking', 'kimble', 'trello']):
            # Project management consolidation
            estimated_savings = int(spend * savings_rules['Consolidate - Project Management']['rate'])
            savings_category = savings_rules['Consolidate - Project Management']['category']
            savings_potential = savings_rules['Consolidate - Project Management']['potential']

        elif any(kw in desc.lower() for kw in ['consulting', 'advisory']):
            # Consulting consolidation
            estimated_savings = int(spend * savings_rules['Consolidate - Consulting']['rate'])
            savings_category = savings_rules['Consolidate - Consulting']['category']
            savings_potential = savings_rules['Consolidate - Consulting']['potential']

        elif any(kw in desc.lower() for kw in ['recruit', 'staffing']):
            # Recruitment consolidation
            estimated_savings = int(spend * savings_rules['Consolidate - Recruitment']['rate'])
            savings_category = savings_rules['Consolidate - Recruitment']['category']
            savings_potential = savings_rules['Consolidate - Recruitment']['potential']

        else:
            # Generic consolidation
            estimated_savings = int(spend * savings_rules['Consolidate - Generic']['rate'])
            savings_category = savings_rules['Consolidate - Generic']['category']
            savings_potential = savings_rules['Consolidate - Generic']['potential']

    elif rec == 'Optimize':
        # Optimization savings based on spend level
        if spend > 100000:
            # High-spend negotiation
            estimated_savings = int(spend * savings_rules['Optimize - High Spend']['rate'])
            savings_category = savings_rules['Optimize - High Spend']['category']
            savings_potential = savings_rules['Optimize - High Spend']['potential']
        elif spend > 10000:
            # Mid-spend optimization
            estimated_savings = int(spend * savings_rules['Optimize - Mid Spend']['rate'])
            savings_category = savings_rules['Optimize - Mid Spend']['category']
            savings_potential = savings_rules['Optimize - Mid Spend']['potential']
        else:
            # Low-spend optimization
            estimated_savings = int(spend * savings_rules['Optimize - Low Spend']['rate'])
            savings_category = savings_rules['Optimize - Low Spend']['category']
            savings_potential = savings_rules['Optimize - Low Spend']['potential']

    # Write to columns
    ws.cell(row, 7).value = savings_category
    ws.cell(row, 8).value = estimated_savings
    ws.cell(row, 9).value = savings_potential

    if estimated_savings > 0:
        savings_count += 1
        savings_total += estimated_savings

print(f"\n{'=' * 140}")
print(f"SAVINGS CALCULATION COMPLETE")
print(f"{'=' * 140}")

print(f"\nVendors with savings potential: {savings_count}")
print(f"Total estimated annual savings: ${savings_total:,.0f}")

print(f"\nSavings by Potential Level:")
for row in range(2, ws.max_row + 1):
    potential = ws.cell(row, 9).value
    if potential:
        # Count and sum by potential
        pass

# Save file with workaround for locks
temp_path = "output/output_file_savings.xlsx"
shutil.copy(file_path, temp_path)

wb_temp = openpyxl.load_workbook(temp_path)
ws_temp = wb_temp['Vendor Analysis Assessment']

# Add headers
ws_temp.cell(1, 7).value = "Savings Category"
ws_temp.cell(1, 8).value = "Estimated Annual Savings (USD)"
ws_temp.cell(1, 9).value = "Savings Potential"

# Recalculate and add savings data
for row in range(2, ws_temp.max_row + 1):
    vendor_name = ws_temp.cell(row, 1).value
    dept = ws_temp.cell(row, 2).value
    spend = ws_temp.cell(row, 3).value or 0
    desc = ws_temp.cell(row, 4).value or ""
    rec = ws_temp.cell(row, 5).value or ""
    note = ws_temp.cell(row, 6).value or ""

    if not vendor_name:
        continue

    savings_category = ""
    estimated_savings = 0
    savings_potential = ""

    if rec == 'Terminate':
        estimated_savings = int(spend * savings_rules['Terminate']['rate'])
        savings_category = savings_rules['Terminate']['category']
        savings_potential = savings_rules['Terminate']['potential']
    elif rec == 'Consolidate':
        if 'navan' in vendor_name.lower():
            estimated_savings = int(spend * savings_rules['Consolidate - Duplicate']['rate'])
            savings_category = savings_rules['Consolidate - Duplicate']['category']
            savings_potential = savings_rules['Consolidate - Duplicate']['potential']
        elif any(kw in desc.lower() for kw in ['real estate', 'office', 'property', 'workspace']):
            estimated_savings = int(spend * savings_rules['Consolidate - Real Estate']['rate'])
            savings_category = savings_rules['Consolidate - Real Estate']['category']
            savings_potential = savings_rules['Consolidate - Real Estate']['potential']
        elif any(kw in desc.lower() for kw in ['cloud', 'infrastructure', 'hosting', 'aws', 'azure']):
            estimated_savings = int(spend * savings_rules['Consolidate - Cloud']['rate'])
            savings_category = savings_rules['Consolidate - Cloud']['category']
            savings_potential = savings_rules['Consolidate - Cloud']['potential']
        elif any(kw in desc.lower() for kw in ['insurance', 'coverage', 'benefit']):
            estimated_savings = int(spend * savings_rules['Consolidate - Insurance']['rate'])
            savings_category = savings_rules['Consolidate - Insurance']['category']
            savings_potential = savings_rules['Consolidate - Insurance']['potential']
        elif any(kw in desc.lower() for kw in ['project', 'task', 'tracking', 'kimble', 'trello']):
            estimated_savings = int(spend * savings_rules['Consolidate - Project Management']['rate'])
            savings_category = savings_rules['Consolidate - Project Management']['category']
            savings_potential = savings_rules['Consolidate - Project Management']['potential']
        elif any(kw in desc.lower() for kw in ['consulting', 'advisory']):
            estimated_savings = int(spend * savings_rules['Consolidate - Consulting']['rate'])
            savings_category = savings_rules['Consolidate - Consulting']['category']
            savings_potential = savings_rules['Consolidate - Consulting']['potential']
        elif any(kw in desc.lower() for kw in ['recruit', 'staffing']):
            estimated_savings = int(spend * savings_rules['Consolidate - Recruitment']['rate'])
            savings_category = savings_rules['Consolidate - Recruitment']['category']
            savings_potential = savings_rules['Consolidate - Recruitment']['potential']
        else:
            estimated_savings = int(spend * savings_rules['Consolidate - Generic']['rate'])
            savings_category = savings_rules['Consolidate - Generic']['category']
            savings_potential = savings_rules['Consolidate - Generic']['potential']
    elif rec == 'Optimize':
        if spend > 100000:
            estimated_savings = int(spend * savings_rules['Optimize - High Spend']['rate'])
            savings_category = savings_rules['Optimize - High Spend']['category']
            savings_potential = savings_rules['Optimize - High Spend']['potential']
        elif spend > 10000:
            estimated_savings = int(spend * savings_rules['Optimize - Mid Spend']['rate'])
            savings_category = savings_rules['Optimize - Mid Spend']['category']
            savings_potential = savings_rules['Optimize - Mid Spend']['potential']
        else:
            estimated_savings = int(spend * savings_rules['Optimize - Low Spend']['rate'])
            savings_category = savings_rules['Optimize - Low Spend']['category']
            savings_potential = savings_rules['Optimize - Low Spend']['potential']

    ws_temp.cell(row, 7).value = savings_category
    ws_temp.cell(row, 8).value = estimated_savings
    ws_temp.cell(row, 9).value = savings_potential

wb_temp.save(temp_path)
wb_temp.close()

# Replace original file
time.sleep(0.5)
try:
    with open(temp_path, 'rb') as f:
        data = f.read()
    with open(file_path, 'wb') as f:
        f.write(data)
    os.remove(temp_path)
    print("\n✓ File updated successfully with savings columns")
except PermissionError:
    print(f"\n⚠ Could not replace original file (locked)")
    print(f"Updated file saved to: {temp_path}")
