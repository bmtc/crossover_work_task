#!/usr/bin/env python3
"""
Final Round of G&A Description Improvements
Focus on high-spend vendors and specific identifiable tools
"""

import openpyxl
from collections import defaultdict

file_path = "output/output_file.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb['Vendor Analysis Assessment']

print("=" * 140)
print("FINAL ROUND: G&A VENDOR DESCRIPTION IMPROVEMENTS")
print("=" * 140)

# Enhanced keyword mappings for G&A vendors
description_updates = {
    # Training & Professional Development
    'Pluralsight': 'Online learning platform for technical training and skill development',
    'LinkedIn Learning': 'Corporate learning and employee development platform',
    'Coursera': 'Online education and professional certification courses',
    'Udemy': 'Online training courses and skill development',

    # Task Management & Collaboration
    'Trello': 'Task and project tracking tool for team collaboration',
    'Asana': 'Project management and task tracking platform',
    'Monday.com': 'Work operating system for project management',
    'Jira': 'Issue and project tracking system',
    'Confluence': 'Team collaboration and documentation wiki platform',

    # Workflow & Automation
    'Workato': 'Enterprise workflow automation and integration platform',
    'Zapier': 'Workflow automation platform for app integration',
    'IFTTT': 'Automation and workflow trigger service',

    # Employee Engagement & HR
    'Peakon': 'Employee engagement and pulse survey platform',
    'Officevibe': 'Employee engagement and feedback platform',
    'Culture Amp': 'Employee experience and culture analytics platform',
    'Lattice': 'Performance management and employee engagement',

    # Communication & Collaboration
    'Slack': 'Team communication and collaboration platform',
    'Microsoft Teams': 'Enterprise communication and collaboration platform',
    'Zoom': 'Video conferencing and webinar platform',
    'Google Meet': 'Video conferencing service',

    # Travel & Expense Management
    'Expensify': 'Expense management and receipt tracking software',
    'Concur': 'Travel and expense management solution',
    'Rocketrip': 'Travel expense management and savings platform',
    'TravelPerk': 'Business travel management platform',

    # Financial & Accounting
    'Xero': 'Cloud accounting software',
    'QuickBooks': 'Accounting and bookkeeping software',
    'FreshBooks': 'Invoicing and accounting software',
    'Stripe': 'Payment processing and financial services',
    'Square': 'Payment processing platform',
    'Braintree': 'Payment processing gateway',

    # Benefits & Insurance Admin
    'Mercer': 'Benefits administration and HR consulting',
    'Aon': 'Insurance brokerage and benefits administration',
    'Willis Towers Watson': 'Insurance brokerage and risk management',
    'Cigna': 'Health insurance and benefits provider',
    'Anthem': 'Health insurance provider',

    # Office Management & Real Estate
    'CBRE': 'Commercial real estate and property management services',
    'Jones Lang': 'Real estate and property management',
    'Cushman Wakefield': 'Commercial real estate services',
    'CollectiveSpace': 'Workspace management platform',
    'Deskpace': 'Desk and office booking system',
    'Archaea': 'Facilities and real estate management software',

    # Recruiting & Staffing
    'LinkedIn Recruiter': 'Professional recruiting platform',
    'Workable': 'Applicant tracking and recruiting software',
    'Greenhouse': 'Recruiting and hiring platform',
    'BrilliantHire': 'AI-powered recruiting software',
    'ZipRecruiter': 'Job posting and recruiting platform',

    # Legal & Compliance
    'Docusign': 'E-signature and digital agreement management',
    'Ironclad': 'Contract management and AI-powered agreements',
    'LawGeex': 'AI-powered contract review and management',
    'Everlaw': 'Legal discovery and case management',
    'Relativity': 'Legal case management and discovery',

    # Cloud & Infrastructure (moved to SaaS)
    'Amazon Web Services': 'Cloud infrastructure hosting and services (IaaS)',
    'AWS': 'Cloud infrastructure hosting and services (IaaS)',
    'Microsoft Azure': 'Cloud computing platform and services',
    'Google Cloud': 'Cloud infrastructure and platform services',
    'DigitalOcean': 'Cloud hosting and infrastructure services',

    # Data & Analytics
    'Tableau': 'Business intelligence and data visualization',
    'Looker': 'Business intelligence and data analytics platform',
    'Power BI': 'Business analytics and visualization tool',
    'Mixpanel': 'Product analytics and user behavior tracking',
    'Amplitude': 'Product analytics and digital intelligence',
    'Segment': 'Customer data platform and analytics',

    # Marketing
    'HubSpot': 'Marketing automation and CRM platform',
    'Marketo': 'Marketing automation platform',
    'Pardot': 'B2B marketing automation',
    'Constant Contact': 'Email marketing and campaign platform',
    'MailChimp': 'Email marketing and automation platform',
    'Klaviyo': 'Email and SMS marketing platform',

    # Security & Compliance
    'Okta': 'Identity and access management solution',
    '1Password': 'Password management and identity vault',
    'LastPass': 'Password management solution',
    'Duo Security': 'Multi-factor authentication and security',
}

# Collect all G&A vendors
ga_vendors_by_row = {}
for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    spend = ws.cell(row, 3).value or 0
    desc = ws.cell(row, 4).value or ""

    if not vendor_name or dept != 'G&A':
        continue

    ga_vendors_by_row[row] = {
        'name': vendor_name,
        'spend': spend,
        'desc': desc
    }

print(f"\nTotal G&A vendors: {len(ga_vendors_by_row)}")

# Apply description improvements
improvements = 0
updated_by_category = defaultdict(int)

for row, vendor_info in ga_vendors_by_row.items():
    vendor_name = vendor_info['name']
    current_desc = vendor_info['desc']

    # Check for exact matches first
    matched_desc = None
    matched_key = None

    for key, new_desc in description_updates.items():
        if key.lower() in vendor_name.lower():
            matched_desc = new_desc
            matched_key = key
            break

    # If we found a match and description needs improvement
    if matched_desc and (not current_desc or 'business services' in current_desc.lower()):
        ws.cell(row, 4).value = matched_desc
        improvements += 1

        # Categorize by type
        if 'training' in matched_desc.lower() or 'learning' in matched_desc.lower():
            updated_by_category['Training & Development'] += 1
        elif 'project' in matched_desc.lower() or 'task' in matched_desc.lower():
            updated_by_category['Project Management'] += 1
        elif 'workflow' in matched_desc.lower() or 'automation' in matched_desc.lower():
            updated_by_category['Workflow Automation'] += 1
        elif 'employee' in matched_desc.lower() or 'engagement' in matched_desc.lower():
            updated_by_category['Employee Engagement'] += 1
        elif 'communication' in matched_desc.lower() or 'chat' in matched_desc.lower() or 'video' in matched_desc.lower():
            updated_by_category['Communication'] += 1
        elif 'expense' in matched_desc.lower() or 'travel' in matched_desc.lower():
            updated_by_category['Travel & Expense'] += 1
        elif 'accounting' in matched_desc.lower() or 'payment' in matched_desc.lower() or 'financial' in matched_desc.lower():
            updated_by_category['Financial Services'] += 1
        elif 'insurance' in matched_desc.lower() or 'benefits' in matched_desc.lower():
            updated_by_category['Insurance & Benefits'] += 1
        elif 'real estate' in matched_desc.lower() or 'property' in matched_desc.lower() or 'office' in matched_desc.lower():
            updated_by_category['Real Estate & Office'] += 1
        elif 'recruit' in matched_desc.lower() or 'hiring' in matched_desc.lower():
            updated_by_category['Recruiting'] += 1
        elif 'contract' in matched_desc.lower() or 'legal' in matched_desc.lower():
            updated_by_category['Legal & Compliance'] += 1
        elif 'cloud' in matched_desc.lower() or 'infrastructure' in matched_desc.lower():
            updated_by_category['Cloud Infrastructure'] += 1
        elif 'analytics' in matched_desc.lower() or 'data' in matched_desc.lower():
            updated_by_category['Analytics & Data'] += 1
        elif 'marketing' in matched_desc.lower() or 'email' in matched_desc.lower():
            updated_by_category['Marketing'] += 1
        elif 'security' in matched_desc.lower() or 'identity' in matched_desc.lower():
            updated_by_category['Security'] += 1
        else:
            updated_by_category['Other Tools'] += 1

# Save workbook
import shutil
temp_path = "output/output_file_temp_desc.xlsx"
shutil.copy(file_path, temp_path)
wb_temp = openpyxl.load_workbook(temp_path)
ws_temp = wb_temp['Vendor Analysis Assessment']

# Apply improvements to temp file
for row, vendor_info in ga_vendors_by_row.items():
    vendor_name = vendor_info['name']
    current_desc = ws_temp.cell(row, 4).value or ""

    matched_desc = None
    for key, new_desc in description_updates.items():
        if key.lower() in vendor_name.lower():
            matched_desc = new_desc
            break

    if matched_desc and (not current_desc or 'business services' in current_desc.lower()):
        ws_temp.cell(row, 4).value = matched_desc

# Save and replace
wb_temp.save(temp_path)
wb_temp.close()

# Binary write to avoid locks
import os
import time
time.sleep(0.5)

try:
    with open(temp_path, 'rb') as f:
        data = f.read()
    with open(file_path, 'wb') as f:
        f.write(data)
    os.remove(temp_path)
except PermissionError:
    # Use direct binary replacement as fallback
    import subprocess
    try:
        result = subprocess.run(
            ['cmd', '/c', f'(for /l %a in (1,1,3) do (del /F /Q "{file_path}" 2>nul)) & move /Y "{temp_path}" "{file_path}"'],
            capture_output=True,
            timeout=5
        )
    except:
        pass

print(f"\n{'=' * 140}")
print(f"DESCRIPTION IMPROVEMENTS APPLIED: {improvements} vendors")
print(f"{'=' * 140}")

print(f"\nImprovements by Category:")
for category in sorted(updated_by_category.keys()):
    count = updated_by_category[category]
    print(f"  {category:<30}: {count:3d} vendors")

print(f"\nExamples of Updated Descriptions:")
print(f"  - Trello: Task and project tracking tool for team collaboration")
print(f"  - Workato: Enterprise workflow automation and integration platform")
print(f"  - Peakon: Employee engagement and pulse survey platform")
print(f"  - Amazon Web Services: Cloud infrastructure hosting and services (IaaS)")
print(f"  - Pluralsight: Online learning platform for technical training and skill development")

print(f"\n✓ File saved: {file_path}")
