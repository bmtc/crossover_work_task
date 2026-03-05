#!/usr/bin/env python3
"""
Strategic Recommendations Engine
Assigns Terminate/Consolidate/Optimize recommendations with red flags
"""

import openpyxl
from collections import defaultdict

file_path = "output/output_file.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb['Vendor Analysis Assessment']

# Add header for red flags column if needed
if ws.cell(1, 6).value is None:
    ws.cell(1, 6).value = "Red Flags / Notes"

print("=" * 120)
print("STRATEGIC RECOMMENDATIONS ENGINE")
print("=" * 120)

# Step 1: Group vendors by department and function
vendors_by_dept = defaultdict(list)
function_groups = defaultdict(list)

for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    spend = ws.cell(row, 3).value or 0
    description = ws.cell(row, 4).value or ""

    if not vendor_name:
        continue

    vendor_info = {
        'row': row,
        'name': vendor_name,
        'dept': dept,
        'spend': spend,
        'desc': description
    }

    vendors_by_dept[dept].append(vendor_info)

    # Create function groups based on keywords
    desc_lower = description.lower()
    func_key = None

    # Sales/CRM functions
    if dept == 'Sales':
        if any(kw in desc_lower for kw in ['crm', 'pipeline', 'deal']):
            func_key = 'CRM'
        elif any(kw in desc_lower for kw in ['intelligence', 'prospecting']):
            func_key = 'Sales Intelligence'
    # Finance functions
    elif dept == 'Finance':
        if any(kw in desc_lower for kw in ['accounting', 'erp', 'ledger']):
            func_key = 'Accounting/ERP'
        elif any(kw in desc_lower for kw in ['audit', 'advisory']):
            func_key = 'Audit/Advisory'
        elif any(kw in desc_lower for kw in ['planning', 'fpa', 'forecast']):
            func_key = 'FP&A'
    # SaaS/IT functions
    elif dept == 'SaaS':
        if any(kw in desc_lower for kw in ['cloud', 'infrastructure', 'hosting', 'iaas']):
            func_key = 'Cloud Infrastructure'
        elif any(kw in desc_lower for kw in ['software', 'ide', 'development']):
            func_key = 'Development Tools'
        elif any(kw in desc_lower for kw in ['networking', 'telecom']):
            func_key = 'Networking'
    # Professional Services functions
    elif dept == 'Professional Services':
        if any(kw in desc_lower for kw in ['recruitment', 'staffing', 'ats']):
            func_key = 'Recruitment'
        elif any(kw in desc_lower for kw in ['consulting', 'advisory']):
            func_key = 'Consulting'
        elif any(kw in desc_lower for kw in ['training', 'development']):
            func_key = 'Training'
    # Facilities functions
    elif dept == 'Facilities':
        if any(kw in desc_lower for kw in ['real estate', 'property', 'office']):
            func_key = 'Real Estate'
        elif any(kw in desc_lower for kw in ['hotel', 'hospitality', 'lodging']):
            func_key = 'Hospitality'
    # G&A functions
    elif dept == 'G&A':
        if any(kw in desc_lower for kw in ['insurance', 'coverage', 'benefits']):
            func_key = 'Insurance'
        elif any(kw in desc_lower for kw in ['travel', 'expense']):
            func_key = 'Travel/Expense'

    if func_key:
        function_groups[func_key].append(vendor_info)

print("\nStep 1: Identified Function Groups")
print("-" * 120)
for func, vendors in sorted(function_groups.items()):
    print(f"\n{func}:")
    for v in sorted(vendors, key=lambda x: x['spend'], reverse=True):
        print(f"  {v['name']:<50} | ${v['spend']:>10,.0f}")

# Step 2: Detect consolidation opportunities
consolidation_candidates = {}  # func_key -> [(vendor_info, reason)]

for func_key, vendors in function_groups.items():
    if len(vendors) > 1:
        # Multiple vendors in same function
        consolidation_candidates[func_key] = []
        for v in vendors:
            reason = f"Multiple {func_key} vendors exist (consolidation opportunity)"
            consolidation_candidates[func_key].append((v, reason))

print("\n\nStep 2: Consolidation Candidates")
print("-" * 120)
consolidation_count = 0
for func_key, candidates in sorted(consolidation_candidates.items()):
    print(f"\n{func_key} ({len(candidates)} vendors):")
    for vendor_info, reason in sorted(candidates, key=lambda x: x[0]['spend'], reverse=True):
        print(f"  {vendor_info['name']:<50} | ${vendor_info['spend']:>10,.0f}")
    consolidation_count += len(candidates)

print(f"\nTotal consolidation candidates: {consolidation_count}")

# Step 3: Detect termination candidates
termination_candidates = []

for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    spend = ws.cell(row, 3).value or 0
    desc = ws.cell(row, 4).value or ""

    if not vendor_name:
        continue

    reason = None

    # Check for red flags
    desc_lower = desc.lower()

    # Very low spend (<$1K)
    if spend < 1000:
        reason = "Very low spend (<$1K) - possible test/legacy tool"

    # Suspicious patterns in name
    if any(pattern in vendor_name.lower() for pattern in ['test', 'demo', 'trial', 'temp', 'old']):
        reason = "Name suggests test/temporary tool"

    # Vague generic descriptions only for single vendors
    if "business services and operations" in desc_lower and spend < 5000:
        if reason:
            reason += " + Vague description"
        else:
            reason = "Vague description + low spend"

    if reason:
        termination_candidates.append((row, vendor_name, dept, spend, reason))

print("\n\nStep 3: Termination Candidates")
print("-" * 120)
print(f"Potential termination candidates: {len(termination_candidates)}")
if termination_candidates:
    print(f"\n{'Vendor Name':<50} | {'Spend':<12} | {'Reason':<40}")
    print("-" * 120)
    for row, name, dept, spend, reason in sorted(termination_candidates, key=lambda x: x[3])[:15]:
        print(f"{name:<50} | ${spend:>10,.0f} | {reason:<40}")

# Step 4: Assign recommendations to all vendors
recommendations_applied = 0
consolidate_count = 0
terminate_count = 0
optimize_count = 0

for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    spend = ws.cell(row, 3).value or 0
    desc = ws.cell(row, 4).value or ""

    if not vendor_name:
        continue

    recommendation = "Optimize"  # Default
    note = ""

    # Check for consolidation
    is_consolidation = False
    for func_key, candidates in consolidation_candidates.items():
        for vendor_info, _ in candidates:
            if vendor_info['row'] == row:
                is_consolidation = True
                break

    if is_consolidation:
        recommendation = "Consolidate"
        note = "Multiple vendors in same function"
        consolidate_count += 1

    # Check for termination (STRONG evidence only)
    is_termination = False
    for t_row, t_name, t_dept, t_spend, t_reason in termination_candidates:
        if t_row == row and t_spend < 500:  # Only terminate very low spend
            is_termination = True
            note = t_reason
            break

    if is_termination:
        recommendation = "Terminate"
        terminate_count += 1
    elif not is_consolidation:
        optimize_count += 1
        if spend > 1000000:
            note = "High-spend vendor - negotiate rates"
        elif spend < 10000:
            note = "Low-spend vendor - verify usage"

    ws.cell(row, 5).value = recommendation
    ws.cell(row, 6).value = note
    recommendations_applied += 1

# Save
wb.save(file_path)

print("\n\n" + "=" * 120)
print("RECOMMENDATIONS APPLIED")
print("=" * 120)
print(f"Total vendors processed: {recommendations_applied}")
print(f"  - Optimize (default): {optimize_count} vendors")
print(f"  - Consolidate: {consolidate_count} vendors")
print(f"  - Terminate: {terminate_count} vendors")

# Summary by department
print("\n\nRecommendations by Department:")
print("-" * 120)

dept_summary = defaultdict(lambda: {'Optimize': 0, 'Consolidate': 0, 'Terminate': 0})

for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    rec = ws.cell(row, 5).value

    if vendor_name and dept and rec:
        dept_summary[dept][rec] += 1

print(f"{'Department':<20} | {'Optimize':<12} | {'Consolidate':<15} | {'Terminate':<12}")
print("-" * 120)
for dept in sorted(dept_summary.keys()):
    stats = dept_summary[dept]
    print(f"{dept:<20} | {stats['Optimize']:<12} | {stats['Consolidate']:<15} | {stats['Terminate']:<12}")

print(f"\n✓ File saved: {file_path}")
