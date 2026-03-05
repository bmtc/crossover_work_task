#!/usr/bin/env python3
"""
Vendor Reclassification using Config Tab Departments ONLY
"""

import openpyxl
import shutil

# Copy input to temp location
shutil.copy("input/input_file.xlsx", "output/output_file_new.xlsx")

file_path = "output/output_file_new.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb['Vendor Analysis Assessment']

CONFIG_DEPARTMENTS = [
    'Engineering', 'Facilities', 'G&A', 'Legal', 'M&A',
    'Marketing', 'SaaS', 'Product', 'Professional Services',
    'Sales', 'Support', 'Finance'
]

# Specific vendor mappings (highest priority)
specific_mappings = {
    'Salesforce Uk Ltd-Uk': ('Sales', 'CRM and sales pipeline management platform'),
    'Navan (Tripactions Inc)': ('G&A', 'Expense management and travel software'),
    'Navan, Inc': ('G&A', 'Expense management and travel software'),
    'Bdo Llp': ('Finance', 'Audit and financial advisory services'),
    'Tog Uk Properties Limited': ('Facilities', 'Office real estate and facilities management'),
    'Cloudcrossing Bvba': ('SaaS', 'Cloud infrastructure and technology services'),
    'Zagrebtower D.O.O.': ('Facilities', 'Office space and facilities management'),
    'Innovent Spaces Private Limited': ('Facilities', 'Office space leasing and design services'),
    'Weking D.O.O.': ('Facilities', 'Office space and business services'),
    'Jensten Insurance Brokers': ('G&A', 'Insurance brokerage and coverage services'),
    'Gpt Space & Co': ('Facilities', 'Office space and facilities management'),
    'Aetna Life And Casualty Ltd': ('G&A', 'Insurance coverage and health benefits'),
    'Rsm Uk Corporate Finance Llp': ('Finance', 'Corporate finance and accounting advisory'),
    'Amazon Web Services Llc': ('SaaS', 'Cloud infrastructure hosting (IaaS)'),
    'Telefonica Global Services Gmbh': ('SaaS', 'Telecommunications and IT services'),
    'Hr Solution International Gmbh': ('Professional Services', 'HR solutions and workforce management'),
    '4I Advisory Services': ('Professional Services', 'Management consulting and advisory'),
    'Bisley Law Ltd': ('Legal', 'Legal counsel and corporate law services'),
    'Infosys': ('SaaS', 'IT services and software development'),
    'Big Frontier Pty Ltd (Cult Of Monday)': ('Product', 'Project management and collaboration platform'),
    'Harmonic Group Limited': ('Professional Services', 'Management consulting'),
    'Wework Singapore Pte. Ltd.': ('Facilities', 'Flexible office space and coworking'),
    'Cloud Technology Solutions Ltd': ('SaaS', 'Cloud infrastructure and technology solutions'),
    'Tmforum': ('SaaS', 'Telecommunications software and standards'),
    'Linkedin Ireland Limited': ('Marketing', 'Social media and professional networking platform'),
    'Kimble Applications Ltd': ('SaaS', 'Project and resource management software'),
    'Sage Uk Limited': ('Finance', 'Accounting and ERP software'),
    'Grant Thornton': ('Finance', 'Audit and financial consulting'),
    'Ss&C Intralinks Inc': ('Finance', 'Financial software and data services'),
    'Veniture D.O.O.': ('G&A', 'Office and business services'),
    'Accutrainee Limited': ('Professional Services', 'Training and talent development services'),
    'Mason Frank International Ltd': ('Professional Services', 'Recruitment and staffing services'),
    'Houlihan Lokey Advisors, Llc': ('Professional Services', 'M&A and financial advisory services'),
    'Vector Capital Management Lp': ('Professional Services', 'Investment and management consulting'),
    'Hubspot Ireland Limited': ('Sales', 'Marketing automation and CRM platform'),
    'Nefron - Obrt Za Poslovne Usluge': ('G&A', 'Business services and operations support'),
    'Planful, Inc.': ('Finance', 'Financial planning and analysis software'),
    'Cognism Limited': ('Sales', 'Sales intelligence and prospecting platform'),
    'Uberflip': ('Marketing', 'Content marketing and engagement platform'),
    'Agram Life Osiguranje D.O.O.': ('G&A', 'Insurance services'),
    'Google Ireland Limited': ('SaaS', 'Cloud services and data analytics platform'),
    'Zuric I Partneri Odvjetnicko Drustvo D.O.O.': ('Legal', 'Legal counsel and corporate law services'),
    'Care Health Insurance Company Limited': ('G&A', 'Health insurance provider'),
    'New Star Networks(Nsn)': ('SaaS', 'IT infrastructure and networking services'),
    'Bupa- Supplier': ('G&A', 'Health insurance and benefits services'),
    'Shree Info System Solutions Pvt Ltd': ('SaaS', 'IT services and software solutions'),
    'Technet It Recruitment': ('Professional Services', 'IT staffing and recruitment services'),
    'Mightyhive Ltd': ('Marketing', 'Digital marketing and advertising services'),
    'Cedar Recruitment Ltd': ('Professional Services', 'Executive recruitment and staffing'),
}

# Keyword-based rules (second priority)
keyword_rules = {
    'Sales': ['salesforce', 'crm', 'hubspot', 'cognism', 'revenue'],
    'Finance': ['accounting', 'audit', 'tax', 'finance', 'bdo', 'grant', 'rsm', 'sage', 'planful', 'intralinks'],
    'SaaS': ['aws', 'cloud', 'infrastructure', 'hosting', 'software', 'platform', 'tech', 'infosys', 'google', 'kimble', 'telefonica', 'development'],
    'Marketing': ['marketing', 'advertising', 'linkedin', 'mightyhive', 'uberflip', 'brand'],
    'Professional Services': ['recruitment', 'recruiting', 'staffing', 'consulting', 'advisory', 'houlihan', 'vector', 'mason', 'accutrainee', 'hr'],
    'Facilities': ['properties', 'real estate', 'office', 'wework', 'space'],
    'G&A': ['insurance', 'travel', 'expense', 'navan', 'bupa', 'aetna', 'brokers', 'care'],
    'Legal': ['law', 'legal', 'counsel', 'solicitor', 'attorney'],
    'Product': ['monday', 'project'],
    'Engineering': ['engineering', 'developer'],
}

print("=" * 100)
print("VENDOR RECLASSIFICATION - CONFIG DEPARTMENTS ONLY")
print("=" * 100)

classified = 0
department_counts = {}

for row in range(2, ws.max_row + 1):
    vendor = ws.cell(row, 1).value
    if not vendor:
        continue

    # First: check specific mappings
    if vendor in specific_mappings:
        dept, desc = specific_mappings[vendor]
    else:
        # Second: check keywords
        vendor_lower = vendor.lower()
        dept = None
        desc = None

        for d, keywords in keyword_rules.items():
            for kw in keywords:
                if kw in vendor_lower:
                    dept = d
                    desc = f"{d} vendor and services"
                    break
            if dept:
                break

        # Default fallback
        if not dept:
            dept = 'G&A'
            desc = 'Business services and operations support'

    # Verify department is valid
    if dept not in CONFIG_DEPARTMENTS:
        print(f"ERROR: Invalid department '{dept}' for vendor '{vendor}'")
        dept = 'G&A'

    ws.cell(row, 2).value = dept
    ws.cell(row, 4).value = desc
    classified += 1
    department_counts[dept] = department_counts.get(dept, 0) + 1

    if classified % 100 == 0:
        print(f"  Processed {classified} vendors...")

print(f"\n✓ Total vendors reclassified: {classified}")
print("\nDepartment Distribution:")

# Calculate by spend
total_by_dept = {}
for row in range(2, ws.max_row + 1):
    vendor = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    spend = ws.cell(row, 3).value or 0
    if vendor and dept:
        total_by_dept[dept] = total_by_dept.get(dept, 0) + spend

total_spend = sum(total_by_dept.values())

for dept in sorted(CONFIG_DEPARTMENTS):
    count = department_counts.get(dept, 0)
    spend = total_by_dept.get(dept, 0)
    pct = (count / classified * 100) if classified > 0 else 0
    spend_pct = (spend / total_spend * 100) if total_spend > 0 else 0
    if count > 0:
        print(f"  {dept:20s}: {count:3d} vendors ({pct:5.1f}%) | ${spend:>13,.2f} ({spend_pct:5.1f}%)")

# Save
wb.save(file_path)
print(f"\n✓ File saved: {file_path}")
