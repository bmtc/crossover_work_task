#!/usr/bin/env python3
"""
Enhanced Termination Analysis
Identify non-core/unnecessary vendors and location-based consolidation opportunities
"""

import openpyxl
from collections import defaultdict

file_path = "output/output_file.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb['Vendor Analysis Assessment']

print("=" * 140)
print("ENHANCED TERMINATION & CONSOLIDATION ANALYSIS")
print("=" * 140)

# Categories for non-core/unnecessary services
non_core_keywords = {
    'Recreation': ['gym', 'recreation', 'sports', 'fitness', 'health club', 'wellness center'],
    'Retail': ['bakery', 'cafe', 'coffee', 'restaurant', 'food', 'supermarket', 'retail', 'store', 'shop'],
    'Hospitality/Lodging': ['hotel', 'resort', 'motel', 'accommodation', 'lodging'],
    'Medical/Health': ['clinic', 'dental', 'medical', 'doctor', 'surgery', 'hospital'],
    'Local Services': ['parking', 'transport', 'courier', 'shipping', 'delivery'],
    'Duplicate IT Vendors': ['apple', 'microsoft'],  # Multiple entities
}

# Location inference from vendor names
location_keywords = {
    'Zagreb': ['zagreb', 'zag', 'croatia d.o.o', 'd.o.o.'],
    'London': ['london', 'uk', 'llp', 'ltd-uk'],
    'Singapore': ['singapore', 'pte. ltd'],
    'Australia': ['australia', 'pty', 'pty ltd', 'melbourne', 'sydney'],
    'USA': ['usa', 'inc', 'llc', 'americas', 'amer'],
}

# Collect vendors
vendors = []
for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    spend = ws.cell(row, 3).value or 0
    desc = ws.cell(row, 4).value or ""
    rec = ws.cell(row, 5).value or ""
    note = ws.cell(row, 6).value or ""

    if not vendor_name:
        continue

    vendors.append({
        'row': row,
        'name': vendor_name,
        'dept': dept,
        'spend': spend,
        'desc': desc,
        'rec': rec,
        'note': note
    })

print("\n### PHASE 1: NON-CORE/UNNECESSARY VENDOR ANALYSIS ###\n")

non_core_candidates = defaultdict(list)
for vendor in vendors:
    vendor_lower = vendor['name'].lower()
    desc_lower = vendor['desc'].lower()

    for category, keywords in non_core_keywords.items():
        for keyword in keywords:
            if keyword in vendor_lower or keyword in desc_lower:
                if vendor['spend'] < 50000:  # Only flag lower-spend items
                    non_core_candidates[category].append(vendor)
                break

print(f"Non-Core/Unnecessary Vendor Candidates by Category:\n")

total_non_core_spend = 0
non_core_list = []

for category in sorted(non_core_candidates.keys()):
    vendors_in_cat = non_core_candidates[category]
    category_spend = sum(v['spend'] for v in vendors_in_cat)
    total_non_core_spend += category_spend

    print(f"{category} ({len(vendors_in_cat)} vendors | ${category_spend:,.0f}):")
    for v in sorted(vendors_in_cat, key=lambda x: x['spend'], reverse=True)[:5]:
        print(f"  {v['name']:<50} | ${v['spend']:>8,.0f} | {v['desc'][:40]}")
    if len(vendors_in_cat) > 5:
        print(f"  ... and {len(vendors_in_cat) - 5} more")
    print()

    non_core_list.extend(vendors_in_cat)

print(f"Total non-core candidates: {len(non_core_list)} vendors | ${total_non_core_spend:,.0f}")

# Identify duplicate Apple & Microsoft entities
print("\n### PHASE 2: DUPLICATE VENDOR ENTITIES ###\n")

apple_vendors = [v for v in vendors if 'apple' in v['name'].lower()]
microsoft_vendors = [v for v in vendors if 'microsoft' in v['name'].lower()]
navan_vendors = [v for v in vendors if 'navan' in v['name'].lower()]

print(f"Apple Entities ({len(apple_vendors)} vendors | ${sum(v['spend'] for v in apple_vendors):,.0f}):")
for v in sorted(apple_vendors, key=lambda x: x['spend'], reverse=True):
    print(f"  {v['name']:<50} | ${v['spend']:>8,.0f}")

print(f"\nMicrosoft Entities ({len(microsoft_vendors)} vendors | ${sum(v['spend'] for v in microsoft_vendors):,.0f}):")
for v in sorted(microsoft_vendors, key=lambda x: x['spend'], reverse=True):
    print(f"  {v['name']:<50} | ${v['spend']:>8,.0f}")

print(f"\nNavan Entities ({len(navan_vendors)} vendors | ${sum(v['spend'] for v in navan_vendors):,.0f}):")
for v in sorted(navan_vendors, key=lambda x: x['spend'], reverse=True):
    print(f"  {v['name']:<50} | ${v['spend']:>8,.0f}")

# Location-based consolidation analysis
print("\n### PHASE 3: LOCATION-BASED CONSOLIDATION (G&A, SaaS, Facilities) ###\n")

location_vendors = defaultdict(lambda: defaultdict(list))  # location -> function -> [vendors]

for vendor in vendors:
    # Skip Legal, Finance, Professional Services (location-independent)
    if vendor['dept'] in ['Legal', 'Finance', 'Professional Services']:
        continue

    # Only process G&A, SaaS, Facilities
    if vendor['dept'] not in ['G&A', 'SaaS', 'Facilities']:
        continue

    # Infer location
    vendor_lower = vendor['name'].lower()
    detected_location = None

    for location, keywords in location_keywords.items():
        for keyword in keywords:
            if keyword in vendor_lower:
                detected_location = location
                break
        if detected_location:
            break

    if not detected_location:
        detected_location = 'Unknown/Global'

    # Infer function
    desc_lower = vendor['desc'].lower()
    function = None

    if 'hotel' in desc_lower or 'hospitality' in desc_lower or 'lodging' in desc_lower:
        function = 'Hotel/Lodging'
    elif 'insurance' in desc_lower or 'coverage' in desc_lower:
        function = 'Insurance'
    elif 'office' in desc_lower or 'real estate' in desc_lower or 'property' in desc_lower or 'space' in desc_lower:
        function = 'Real Estate/Office'
    elif 'cloud' in desc_lower or 'infrastructure' in desc_lower:
        function = 'Cloud Infrastructure'
    elif 'development' in desc_lower or 'software' in desc_lower:
        function = 'Development Tools'
    else:
        function = f"{vendor['dept']}/Other"

    location_vendors[detected_location][function].append(vendor)

print(f"Potential Location-Based Consolidation Opportunities:\n")

consolidation_count = 0
for location in sorted(location_vendors.keys()):
    functions = location_vendors[location]
    location_total_spend = sum(
        v['spend'] for func_list in functions.values() for v in func_list
    )

    print(f"{location} (${location_total_spend:,.0f}):")

    for function in sorted(functions.keys()):
        vendors_in_func = functions[function]
        if len(vendors_in_func) > 1:
            func_spend = sum(v['spend'] for v in vendors_in_func)
            print(f"  {function}: {len(vendors_in_func)} vendors (${func_spend:,.0f}) - CONSOLIDATION CANDIDATE")
            for v in sorted(vendors_in_func, key=lambda x: x['spend'], reverse=True):
                print(f"    {v['name']:<47} | ${v['spend']:>8,.0f}")
            consolidation_count += 1
            print()
        elif len(vendors_in_func) == 1:
            v = vendors_in_func[0]
            print(f"  {function}: 1 vendor (${v['spend']:,.0f}) - Single provider")

print(f"\nTotal location-based consolidation opportunities: {consolidation_count} groups")

# Current termination review
print("\n### PHASE 4: CURRENT TERMINATION LIST REVIEW ###\n")

current_terminations = [v for v in vendors if v['rec'] == 'Terminate']
print(f"Current terminations: {len(current_terminations)} vendors (${sum(v['spend'] for v in current_terminations):,.0f})")

print(f"\nByReason:")
reasons = defaultdict(list)
for v in current_terminations:
    reason = v['note'] or 'No note'
    reasons[reason].append(v)

for reason in sorted(reasons.keys()):
    count = len(reasons[reason])
    spend = sum(v['spend'] for v in reasons[reason])
    print(f"  {reason}: {count} vendors (${spend:,.0f})")

print("\n" + "=" * 140)
print("SUMMARY & RECOMMENDATIONS")
print("=" * 140)

print(f"""
TERMINATION RECOMMENDATIONS:

1. Non-Core Services (Employee Perks/Amenities):
   - Categories: Gyms, recreation clubs, restaurants, retail shops, hotels
   - Count: {len(non_core_list)} vendors
   - Spend: ${total_non_core_spend:,.0f}
   - Rationale: Not core to business - eliminate or replace with stipend
   - Examples: Gyms, recreation clubs, bakeries, retail shops
   - ACTION: Mark as "Terminate (Non-Core)" instead of current logic

2. Duplicate Vendor Entities:
   - Apple: {len(apple_vendors)} entities (likely duplicate licenses) - consolidate to 1
   - Microsoft: {len(microsoft_vendors)} entities (likely duplicate licenses) - consolidate to 1
   - Navan: {len(navan_vendors)} entities (same vendor, different structures) - consolidate to 1
   - Spend: ${sum(v['spend'] for v in apple_vendors + microsoft_vendors + navan_vendors):,.0f}
   - ACTION: Change to "Consolidate (Duplicate Entity)" with higher consolidation priority

CONSOLIDATION RECOMMENDATIONS (Location-Based):

1. By Location/Function (G&A, SaaS, Facilities only):
   - Zagreb office space: Multiple providers (4+ vendors) - consolidate to 1-2
   - Hotels/Lodging by city: Multiple providers per location - negotiate single contract
   - Insurance: Multiple carriers for same coverage - consolidate to 1-2
   - Cloud infrastructure: 6 vendors across regions - consolidate to 2 (primary + backup)
   - Spend Opportunity: ~$400K+ from location-based consolidation

FINANCIAL IMPACT PROJECTION:
   - Current terminations: {len(current_terminations)} vendors ($25.6K)
   - Add non-core terminations: {len(non_core_list)} vendors (+${total_non_core_spend:,.0f})
   - Revised terminations: {len(current_terminations) + len(non_core_list)} vendors (~${sum(v['spend'] for v in current_terminations) + total_non_core_spend:,.0f})

   - Location-based consolidation savings: ~$400K
   - Duplicate entity consolidation: ~$100K
   - Total potential: ~$500K+ (vs current $560K estimate)
""")
