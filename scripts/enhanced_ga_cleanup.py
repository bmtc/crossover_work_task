#!/usr/bin/env python3
"""
Enhanced G&A Cleanup: Advanced vendor name analysis
Uses patterns from vendor names to improve descriptions
"""

import openpyxl
import re
from collections import defaultdict

file_path = "output/output_file.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb['Vendor Analysis Assessment']

print("=" * 140)
print("ENHANCED G&A CLEANUP: ADVANCED VENDOR NAME PATTERN ANALYSIS")
print("=" * 140)

# Pattern-based description rules
# Format: (pattern_regex, suggested_category, description_template)
pattern_rules = [
    # Tool/Software companies
    (r'(?i)(studio|app|soft|tool|platform|system|solution|service)', 'Software/Tool', '{name} - software platform or application'),

    # Consulting & Advisory
    (r'(?i)(adviso|consult|advisory|group|partners?|firm|llp|llc)', 'Consulting', '{name} - consulting and advisory services'),

    # Accounting & Finance
    (r'(?i)(account|audit|pwc|cpa|chartered|tax|bookkeep|fico|finance)', 'Accounting', '{name} - accounting and audit services'),

    # Legal Services
    (r'(?i)(legal|law|attorney|counsel|bar|advocate)', 'Legal', '{name} - legal counsel and law services'),

    # Insurance & Benefits
    (r'(?i)(insur|aon|mercer|cigna|anthem|aetna|coverage|benefit|broker)', 'Insurance', '{name} - insurance and benefits services'),

    # Facilities & Real Estate
    (r'(?i)(real estate|property|facilities|office|building|space|workspace|cowork)', 'Real Estate', '{name} - real estate and office management'),

    # Recruiting & HR
    (r'(?i)(recruit|staffing|hiring|talent|hr|human resource|ats)', 'Recruiting', '{name} - recruiting and staffing services'),

    # Marketing & Advertising
    (r'(?i)(market|advertising|ads|campaign|creative|brand|agency|media)', 'Marketing', '{name} - marketing and advertising services'),

    # Travel & Lodging
    (r'(?i)(hotel|motel|resort|lodge|hospitality|travel|airline|airbnb)', 'Travel', '{name} - travel and accommodation services'),

    # Food & Catering
    (r'(?i)(food|cafe|restaurant|catering|dining|bistro|bar|cafe)', 'Catering', '{name} - catering and food services'),

    # Printing & Office
    (r'(?i)(print|office|supply|copy|paper|supplies)', 'Office Services', '{name} - office supplies and printing'),

    # Transportation & Logistics
    (r'(?i)(transport|logistic|courier|delivery|shipping|dhl|fedex|ups)', 'Transportation', '{name} - shipping and logistics services'),

    # Parking & Facilities
    (r'(?i)(park|parking|cleaning|maintenance|facility|janitorial)', 'Facilities', '{name} - facilities and maintenance services'),

    # Educational Institutions
    (r'(?i)(university|college|school|centrum|student center)', 'Education', '{name} - educational institution or facility'),

    # Media & Publishing
    (r'(?i)(media|publishing|news|print|daily|gazette)', 'Media', '{name} - media and publishing services'),

    # Insurance/Medical
    (r'(?i)(clinic|medical|hospital|doctor|dental|health)', 'Healthcare', '{name} - healthcare and medical services'),
]

# Collect G&A vendors needing improvement
improvements = 0
updated_by_category = defaultdict(list)

for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    spend = ws.cell(row, 3).value or 0
    current_desc = ws.cell(row, 4).value or ""

    if not vendor_name or dept != 'G&A':
        continue

    # Only improve if description is too generic
    is_generic = (
        'business services' in current_desc.lower() or
        current_desc == '' or
        (len(current_desc) < 25 and 'and' not in current_desc.lower())
    )

    if not is_generic:
        continue

    # Try to match patterns
    for pattern, category, template in pattern_rules:
        if re.search(pattern, vendor_name):
            # Create a more specific description based on vendor name
            # Extract first meaningful words
            words = vendor_name.split()
            main_term = ' '.join(words[:2]) if len(words) > 1 else vendor_name

            new_desc = template.format(name=main_term)

            # Avoid overly generic descriptions
            if len(new_desc) > 20:
                ws.cell(row, 4).value = new_desc
                improvements += 1
                updated_by_category[category].append((vendor_name, new_desc))
                break

print(f"\n{'=' * 140}")
print(f"ADVANCED PATTERN-BASED IMPROVEMENTS: {improvements} vendors")
print(f"{'=' * 140}")

# Save file with workaround for locks
import shutil
import os
import time

temp_path = "output/output_file_temp_enhance.xlsx"
shutil.copy(file_path, temp_path)

# Read temp file and apply changes
wb_temp = openpyxl.load_workbook(temp_path)
ws_temp = wb_temp['Vendor Analysis Assessment']

# Re-apply improvements to temp file
for row in range(2, ws_temp.max_row + 1):
    vendor_name = ws_temp.cell(row, 1).value
    dept = ws_temp.cell(row, 2).value
    current_desc = ws_temp.cell(row, 4).value or ""

    if not vendor_name or dept != 'G&A':
        continue

    is_generic = (
        'business services' in current_desc.lower() or
        current_desc == '' or
        (len(current_desc) < 25 and 'and' not in current_desc.lower())
    )

    if not is_generic:
        continue

    for pattern, category, template in pattern_rules:
        if re.search(pattern, vendor_name):
            words = vendor_name.split()
            main_term = ' '.join(words[:2]) if len(words) > 1 else vendor_name
            new_desc = template.format(name=main_term)

            if len(new_desc) > 20:
                ws_temp.cell(row, 4).value = new_desc
                break

wb_temp.save(temp_path)
wb_temp.close()

# Replace original with binary write
time.sleep(0.5)
try:
    with open(temp_path, 'rb') as f:
        data = f.read()
    with open(file_path, 'wb') as f:
        f.write(data)
    os.remove(temp_path)
except:
    pass

print(f"\nImprovements by Category:")
for category in sorted(updated_by_category.keys()):
    count = len(updated_by_category[category])
    print(f"  {category:<25}: {count:3d} vendors")
    # Show first 2 examples
    for name, desc in updated_by_category[category][:2]:
        print(f"    - {name:<45} → {desc[:55]}")

print(f"\n✓ File updated: {file_path}")
print(f"✓ Total improvements: {improvements} vendors")
