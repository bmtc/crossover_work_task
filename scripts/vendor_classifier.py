#!/usr/bin/env python3
"""
Vendor Classification Script
Classifies all vendors into departments and generates specific descriptions.
"""

import openpyxl

# Load the output file
file_path = "output/output_file.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb['Vendor Analysis Assessment']

# Define classification rules with keyword mapping
vendor_mappings = {
    'Salesforce Uk Ltd-Uk': ('Sales', 'CRM and sales pipeline management platform'),
    'Navan (Tripactions Inc)': ('G&A/Operations', 'Expense management and travel software'),
    'Navan, Inc': ('G&A/Operations', 'Expense management and travel software'),
    'Bdo Llp': ('Finance', 'Audit and financial advisory services'),
    'Tog Uk Properties Limited': ('G&A/Operations', 'Office real estate and facilities management'),
    'Cloudcrossing Bvba': ('Engineering', 'Cloud infrastructure and technology services'),
    'Zagrebtower D.O.O.': ('G&A/Operations', 'Office space and facilities management'),
    'Innovent Spaces Private Limited': ('G&A/Operations', 'Office space leasing and design services'),
    'Weking D.O.O.': ('G&A/Operations', 'Office space and business services'),
    'Jensten Insurance Brokers': ('G&A/Operations', 'Insurance brokerage and coverage services'),
    'Gpt Space & Co': ('G&A/Operations', 'Office space and facilities management'),
    'Aetna Life And Casualty Ltd': ('G&A/Operations', 'Insurance coverage and health benefits'),
    'Rsm Uk Corporate Finance Llp': ('Finance', 'Corporate finance and accounting advisory'),
    'Amazon Web Services Llc': ('IT', 'Cloud infrastructure hosting (IaaS)'),
    'Telefonica Global Services Gmbh': ('IT', 'Telecommunications and IT services'),
    'Hr Solution International Gmbh': ('People/HR', 'HR solutions and workforce management'),
    '4I Advisory Services': ('Consulting', 'Management consulting and advisory services'),
    'Bisley Law Ltd': ('Legal', 'Legal counsel and corporate law services'),
    'Infosys': ('Engineering', 'IT services and software development'),
    'Big Frontier Pty Ltd (Cult Of Monday)': ('Product', 'Project management and work collaboration platform'),
    'Harmonic Group Limited': ('Consulting', 'Management consulting services'),
    'Wework Singapore Pte. Ltd.': ('G&A/Operations', 'Flexible office space and coworking'),
    'Cloud Technology Solutions Ltd': ('IT', 'Cloud infrastructure and technology solutions'),
    'Tmforum': ('Engineering', 'Telecommunications software and standards'),
    'Linkedin Ireland Limited': ('Marketing', 'Social media and professional networking platform'),
    'Kimble Applications Ltd': ('Engineering', 'Project and resource management software'),
    'Sage Uk Limited': ('Finance', 'Accounting and ERP software'),
    'Grant Thornton': ('Finance', 'Audit and financial consulting'),
    'Ss&C Intralinks Inc': ('Finance', 'Financial software and data services'),
    'Veniture D.O.O.': ('G&A/Operations', 'Office and business services'),
    'Accutrainee Limited': ('People/HR', 'Training and talent development services'),
    'Mason Frank International Ltd': ('People/HR', 'Recruitment and staffing services'),
    'Houlihan Lokey Advisors, Llc': ('Consulting', 'M&A and financial advisory services'),
    'Vector Capital Management Lp': ('Consulting', 'Investment and management consulting'),
    'Hubspot Ireland Limited': ('Sales', 'Marketing automation and CRM platform'),
    'Nefron - Obrt Za Poslovne Usluge': ('G&A/Operations', 'Business services and operations support'),
    'Planful, Inc.': ('Finance', 'Financial planning and analysis software'),
    'Cognism Limited': ('Sales', 'Sales intelligence and prospecting platform'),
    'Uberflip': ('Marketing', 'Content marketing and engagement platform'),
    'Agram Life Osiguranje D.O.O.': ('G&A/Operations', 'Insurance services'),
    'Google Ireland Limited': ('IT', 'Cloud services and data analytics platform'),
    'Zuric I Partneri Odvjetnicko Drustvo D.O.O.': ('Legal', 'Legal counsel and corporate law services'),
    'Care Health Insurance Company Limited': ('G&A/Operations', 'Health insurance provider'),
    'New Star Networks(Nsn)': ('IT', 'IT infrastructure and networking services'),
    'Bupa- Supplier': ('G&A/Operations', 'Health insurance and benefits services'),
    'Shree Info System Solutions Pvt Ltd': ('Engineering', 'IT services and software solutions'),
    'Technet It Recruitment': ('People/HR', 'IT staffing and recruitment services'),
    'Mightyhive Ltd': ('Marketing', 'Digital marketing and advertising services'),
    'Cedar Recruitment Ltd': ('People/HR', 'Executive recruitment and staffing'),
}

def classify_vendor(vendor_name):
    """Classify vendor with smart keyword matching"""
    # Check if we have a direct mapping
    if vendor_name in vendor_mappings:
        return vendor_mappings[vendor_name]

    # Keyword-based classification
    vendor_lower = vendor_name.lower()

    # Finance
    if any(k in vendor_lower for k in ['bank', 'accounting', 'audit', 'tax', 'finance', 'bdo', 'deloitte', 'pwc', 'kpmg', 'ernst', 'grant', 'rsm', 'cpa', 'planful', 'sage', 'ss&c', 'intralinks']):
        if 'tax' in vendor_lower:
            return ('Finance', 'Tax advisory and compliance services')
        if 'audit' in vendor_lower:
            return ('Finance', 'Audit and financial advisory services')
        if 'accounting' in vendor_lower or 'sage' in vendor_lower or 'planful' in vendor_lower:
            return ('Finance', 'Accounting and financial software')
        return ('Finance', 'Financial services and advisory')

    # HR/People
    if any(k in vendor_lower for k in ['recruitment', 'recruiting', 'ats', 'hr', 'human resources', 'talent', 'headhunt', 'staffing']):
        if 'recruitment' in vendor_lower or 'recruiting' in vendor_lower or 'staffing' in vendor_lower:
            return ('People/HR', 'Recruitment and staffing services')
        return ('People/HR', 'HR and talent management services')

    # Legal
    if any(k in vendor_lower for k in ['law', 'legal', 'attorney', 'counsel', 'solicitor']):
        return ('Legal', 'Legal counsel and advisory services')

    # Sales/CRM
    if any(k in vendor_lower for k in ['salesforce', 'crm', 'hubspot', 'cognism', 'revenue']):
        return ('Sales', 'Sales and CRM platform')

    # Marketing
    if any(k in vendor_lower for k in ['marketing', 'advertising', 'uberflip', 'linkedin', 'mightyhive', 'brand']):
        return ('Marketing', 'Marketing and advertising services')

    # IT/Cloud/Infrastructure
    if any(k in vendor_lower for k in ['aws', 'amazon web', 'azure', 'google cloud', 'cloud', 'infrastructure', 'hosting', 'network', 'data center', 'telefonica']):
        if 'cloud' in vendor_lower or 'aws' in vendor_lower or 'azure' in vendor_lower:
            return ('IT', 'Cloud infrastructure and hosting services')
        return ('IT', 'IT infrastructure and operations')

    # Engineering/Software
    if any(k in vendor_lower for k in ['software', 'platform', 'saas', 'development', 'tech', 'infosys', 'cloudcrossing', 'kimble']):
        return ('Engineering', 'Software development and platform services')

    # Product
    if 'monday' in vendor_lower or 'project' in vendor_lower:
        return ('Product', 'Project management and collaboration platform')

    # Consulting
    if any(k in vendor_lower for k in ['advisory', 'consulting', 'strategy', 'advisors', 'houlihan', 'vector']):
        return ('Consulting', 'Management consulting and advisory services')

    # G&A/Operations (default)
    if any(k in vendor_lower for k in ['properties', 'real estate', 'space', 'office', 'wework', 'insurance', 'brokers', 'travel', 'expense', 'navan', 'facilities', 'bupa', 'aetna']):
        if 'insurance' in vendor_lower or 'brokers' in vendor_lower or 'bupa' in vendor_lower or 'aetna' in vendor_lower:
            return ('G&A/Operations', 'Insurance and benefits services')
        if 'space' in vendor_lower or 'properties' in vendor_lower or 'office' in vendor_lower or 'wework' in vendor_lower:
            return ('G&A/Operations', 'Office space and facilities management')
        if 'travel' in vendor_lower or 'expense' in vendor_lower or 'navan' in vendor_lower:
            return ('G&A/Operations', 'Travel and expense management software')
        return ('G&A/Operations', 'Facilities and operations services')

    # Default fallback
    return ('G&A/Operations', 'Business services and operations support')

# Process all vendors
print("=" * 100)
print("VENDOR CLASSIFICATION EXECUTION")
print("=" * 100)

classified_count = 0
department_counts = {}

for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    if vendor_name:
        dept, desc = classify_vendor(vendor_name)
        ws.cell(row, 2).value = dept
        ws.cell(row, 4).value = desc
        classified_count += 1

        # Track department counts
        department_counts[dept] = department_counts.get(dept, 0) + 1

        if classified_count % 50 == 0:
            print(f"  Processed {classified_count} vendors...")

print(f"\n✓ Total vendors classified: {classified_count}")
print("\nDepartment Distribution:")
for dept in sorted(department_counts.keys()):
    count = department_counts[dept]
    pct = (count / classified_count) * 100
    print(f"  {dept:20s}: {count:3d} vendors ({pct:5.1f}%)")

print("\nSample Classifications (first 15):")
print(f"{'Vendor Name':<45} | {'Department':<20} | {'Description':<40}")
print("-" * 110)
for row in range(2, 17):
    vendor = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    desc = ws.cell(row, 4).value
    if vendor:
        print(f"{vendor:<45} | {dept:<20} | {desc:<40}")

# Save
wb.save(file_path)
print(f"\n✓ File saved: {file_path}")
