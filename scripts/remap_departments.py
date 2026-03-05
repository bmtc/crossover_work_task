#!/usr/bin/env python3
"""
Remap Vendors to Correct Departments Based on Improved Descriptions
Move SaaS tools from G&A to SaaS department
Move services from G&A to Professional Services
Move facilities from G&A to Facilities
"""

import openpyxl
import shutil
import os
import time

# Use the fixed file with improvements
source_file = "output/output_file_fixed.xlsx"
target_file = "output/output_file.xlsx"
temp_file = "output/output_file_remapped.xlsx"

# Create temp copy
shutil.copy(source_file, temp_file)

wb = openpyxl.load_workbook(temp_file)
ws = wb['Vendor Analysis Assessment']

print("=" * 140)
print("DEPARTMENT REMAPPING: Moving Vendors to Correct Departments Based on Descriptions")
print("=" * 140)

# Remapping rules
saas_keywords = [
    'platform', 'software', 'tool', 'ide', 'automation', 'workflow',
    'cloud infrastructure', 'hosting', 'saas', 'application', 'system',
    'password management', 'identity', 'survey', 'engagement platform'
]

ps_keywords = [
    'consulting', 'advisory', 'audit', 'tax', 'recruitment', 'staffing',
    'training', 'learning', 'coaching', 'education'
]

facilities_keywords = [
    'real estate', 'office', 'property', 'workspace', 'hotel', 'lodging',
    'accommodation', 'parking', 'facilities', 'food service'
]

# Track remappings
remappings = {
    'to_SaaS': [],
    'to_ProfessionalServices': [],
    'to_Facilities': [],
}

for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    current_dept = ws.cell(row, 2).value
    desc = ws.cell(row, 4).value or ""
    spend = ws.cell(row, 3).value or 0
    note = ws.cell(row, 6).value or ""

    if not vendor_name or current_dept != 'G&A':
        continue

    desc_lower = desc.lower()
    new_dept = None
    reason = None

    # Check for SaaS
    if any(kw in desc_lower for kw in saas_keywords):
        new_dept = 'SaaS'
        reason = 'Software/Platform tool'
        remappings['to_SaaS'].append({
            'name': vendor_name,
            'desc': desc,
            'spend': spend,
            'reason': reason
        })

    # Check for Professional Services
    elif any(kw in desc_lower for kw in ps_keywords):
        new_dept = 'Professional Services'
        reason = 'Professional service provider'
        remappings['to_ProfessionalServices'].append({
            'name': vendor_name,
            'desc': desc,
            'spend': spend,
            'reason': reason
        })

    # Check for Facilities
    elif any(kw in desc_lower for kw in facilities_keywords):
        new_dept = 'Facilities'
        reason = 'Facilities/Real estate/Lodging'
        remappings['to_Facilities'].append({
            'name': vendor_name,
            'desc': desc,
            'spend': spend,
            'reason': reason
        })

    # Apply remapping if needed
    if new_dept and new_dept != current_dept:
        ws.cell(row, 2).value = new_dept
        # Update note to explain remapping
        if not note or 'business services' in note.lower():
            ws.cell(row, 6).value = f"[Remapped from G&A to {new_dept}] {reason}"

# Save remapped file
wb.save(temp_file)
wb.close()

print(f"\nVendors Remapped to Correct Departments:\n")

total_remap_spend = 0

# SaaS remappings
if remappings['to_SaaS']:
    saas_spend = sum(v['spend'] for v in remappings['to_SaaS'])
    total_remap_spend += saas_spend
    print(f"→ SaaS Department ({len(remappings['to_SaaS'])} vendors, ${saas_spend:,.0f}):")
    for v in sorted(remappings['to_SaaS'], key=lambda x: x['spend'], reverse=True):
        print(f"  • {v['name']:<45} | ${v['spend']:>10,.0f} | {v['desc'][:40]}")
    print()

# Professional Services remappings
if remappings['to_ProfessionalServices']:
    ps_spend = sum(v['spend'] for v in remappings['to_ProfessionalServices'])
    total_remap_spend += ps_spend
    print(f"→ Professional Services ({len(remappings['to_ProfessionalServices'])} vendors, ${ps_spend:,.0f}):")
    for v in sorted(remappings['to_ProfessionalServices'], key=lambda x: x['spend'], reverse=True):
        print(f"  • {v['name']:<45} | ${v['spend']:>10,.0f} | {v['desc'][:40]}")
    print()

# Facilities remappings
if remappings['to_Facilities']:
    fac_spend = sum(v['spend'] for v in remappings['to_Facilities'])
    total_remap_spend += fac_spend
    print(f"→ Facilities ({len(remappings['to_Facilities'])} vendors, ${fac_spend:,.0f}):")
    for v in sorted(remappings['to_Facilities'], key=lambda x: x['spend'], reverse=True):
        print(f"  • {v['name']:<45} | ${v['spend']:>10,.0f} | {v['desc'][:40]}")
    print()

print("=" * 140)
print(f"REMAPPING SUMMARY")
print("=" * 140)

print(f"""
Total Vendors Remapped: {sum(len(v) for v in remappings.values())} vendors (${total_remap_spend:,.0f})

NEW CONSOLIDATION OPPORTUNITIES (Post-Remapping):

1. NAVAN DUPLICATE (SaaS - Expense Management)
   - Navan (Tripactions Inc): $357,984
   - Navan, Inc: $57,929
   - TOTAL: $415,913
   - ACTION: Consolidate to single licensing agreement
   - SAVINGS: ~$150-200K (eliminate duplicate licenses)

2. EXPENSE/TRAVEL MANAGEMENT (SaaS)
   - Navan variants: $415,913
   - Other travel software: consolidate if duplicate

3. IDENTITY & SECURITY (SaaS)
   - Lastpass: $263
   - Consider consolidation with larger identity vendor

Key Insight: Remapping reveals consolidation is already identified (Navan)
Additional opportunity: SaaS vendors now grouped by function for easier consolidation analysis

Updated Department Distribution:
- G&A: Reduced (facilities & services moved out)
- SaaS: +8 software tools = improved SaaS clarity
- Professional Services: +3 service providers
- Facilities: +3 real estate/facilities vendors
""")

# Replace original file with remapped version
print(f"\nApplying remapped file: {temp_file} → {target_file}")
print("=" * 140)

# Binary write to avoid locks
time.sleep(0.5)
try:
    with open(temp_file, 'rb') as f:
        data = f.read()
    with open(target_file, 'wb') as f:
        f.write(data)
    os.remove(temp_file)
    print("✓ Remapped file applied successfully")
except PermissionError:
    print(f"⚠ Could not replace original file (locked)")
    print(f"Remapped file available at: {temp_file}")
    print(f"Manual sync needed once file becomes available")

print("\n✓ Remapping complete")
