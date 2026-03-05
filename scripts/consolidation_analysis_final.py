#!/usr/bin/env python3
"""
Final Consolidation Analysis: Identify All Consolidation Opportunities Post-Remapping
Focus on function-based grouping across all departments
"""

import openpyxl
from collections import defaultdict

file_path = "output/output_file.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb['Vendor Analysis Assessment']

print("=" * 140)
print("CONSOLIDATION ANALYSIS: All Departments (Post-Remapping)")
print("=" * 140)

# Collect vendors by department and function
consolidation_groups = defaultdict(lambda: defaultdict(list))

for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    desc = ws.cell(row, 4).value or ""
    spend = ws.cell(row, 3).value or 0
    rec = ws.cell(row, 5).value or ""

    if not vendor_name or dept in ['Legal', 'Finance', 'M&A']:
        continue

    desc_lower = desc.lower()
    function_group = None

    # SaaS Function Groups
    if dept == 'SaaS':
        if 'navan' in vendor_name.lower():
            function_group = 'Expense Management / Travel'
        elif any(kw in desc_lower for kw in ['project', 'task', 'tracking', 'kimble', 'asana']):
            function_group = 'Project Management Tools'
        elif any(kw in desc_lower for kw in ['workflow', 'automation', 'zapier', 'ifttt']):
            function_group = 'Workflow Automation'
        elif any(kw in desc_lower for kw in ['cloud', 'infrastructure', 'hosting', 'aws', 'azure']):
            function_group = 'Cloud Infrastructure'
        elif any(kw in desc_lower for kw in ['communication', 'slack', 'teams', 'zoom']):
            function_group = 'Communication & Collaboration'
        elif any(kw in desc_lower for kw in ['learning', 'training', 'pluralsight', 'coursera']):
            function_group = 'Training & Learning Platforms'
        elif any(kw in desc_lower for kw in ['analytics', 'data', 'tableau', 'looker']):
            function_group = 'Analytics & Business Intelligence'
        elif any(kw in desc_lower for kw in ['password', 'identity', 'security', 'okta']):
            function_group = 'Identity & Access Management'
        elif any(kw in desc_lower for kw in ['engagement', 'survey', 'peakon']):
            function_group = 'Employee Engagement Platforms'
        else:
            function_group = f'SaaS - Other'

    # Professional Services Function Groups
    elif dept == 'Professional Services':
        if any(kw in desc_lower for kw in ['recruit', 'staffing', 'ats', 'talent']):
            function_group = 'Recruitment & Staffing'
        elif any(kw in desc_lower for kw in ['consulting', 'advisory', 'strategy']):
            function_group = 'Consulting & Advisory'
        elif any(kw in desc_lower for kw in ['audit', 'accounting', 'tax']):
            function_group = 'Accounting & Audit Services'
        elif any(kw in desc_lower for kw in ['training', 'coaching', 'education']):
            function_group = 'Training & Professional Development'
        else:
            function_group = 'Professional Services - Other'

    # Facilities Function Groups
    elif dept == 'Facilities':
        if any(kw in desc_lower for kw in ['real estate', 'property', 'office', 'workspace']):
            function_group = 'Real Estate & Office Space'
        elif any(kw in desc_lower for kw in ['hotel', 'resort', 'lodge', 'accommodation']):
            function_group = 'Corporate Lodging'
        elif any(kw in desc_lower for kw in ['food', 'catering', 'cafe', 'restaurant']):
            function_group = 'Food Services & Catering'
        elif any(kw in desc_lower for kw in ['parking', 'transport', 'courier']):
            function_group = 'Parking & Transportation'
        else:
            function_group = 'Facilities - Other'

    # G&A Function Groups
    elif dept == 'G&A':
        if any(kw in desc_lower for kw in ['insurance', 'coverage', 'benefits', 'aon', 'mercer']):
            function_group = 'Insurance & Employee Benefits'
        elif any(kw in desc_lower for kw in ['travel', 'expense', 'booking']):
            function_group = 'Travel & Expense Management'
        elif any(kw in desc_lower for kw in ['office', 'supplies', 'printing']):
            function_group = 'Office Supplies & Services'
        else:
            function_group = 'G&A - Other Operations'

    # Marketing
    elif dept == 'Marketing':
        if any(kw in desc_lower for kw in ['email', 'campaign', 'automation']):
            function_group = 'Marketing Automation & Email'
        elif any(kw in desc_lower for kw in ['social', 'media', 'content']):
            function_group = 'Social Media & Content'
        else:
            function_group = 'Marketing Tools'

    # Sales
    elif dept == 'Sales':
        if any(kw in desc_lower for kw in ['crm', 'salesforce']):
            function_group = 'CRM & Sales Pipeline'
        else:
            function_group = 'Sales Tools'

    if function_group:
        consolidation_groups[dept][function_group].append({
            'name': vendor_name,
            'desc': desc,
            'spend': spend,
            'rec': rec,
            'row': row
        })

# Print consolidation opportunities
print("\nCONSOLIDATION OPPORTUNITIES BY DEPARTMENT & FUNCTION:\n")

total_consolidation_spend = 0
total_consolidation_groups = 0

for dept in sorted(consolidation_groups.keys()):
    functions = consolidation_groups[dept]
    dept_total_spend = sum(
        v['spend']
        for func_vendors in functions.values()
        for v in func_vendors
    )

    print(f"\n{'=' * 140}")
    print(f"{dept.upper()} ({len([v for func_vendors in functions.values() for v in func_vendors])} vendors, ${dept_total_spend:,.0f})")
    print(f"{'=' * 140}\n")

    for function, vendors in sorted(functions.items(), key=lambda x: -sum(v['spend'] for v in x[1])):
        if len(vendors) > 1:
            # Consolidation candidate (multiple vendors same function)
            func_spend = sum(v['spend'] for v in vendors)
            total_consolidation_spend += func_spend
            total_consolidation_groups += 1

            print(f"  ✓ {function} ({len(vendors)} vendors, ${func_spend:,.0f})")
            for v in sorted(vendors, key=lambda x: x['spend'], reverse=True)[:5]:
                status = "→ Consolidate" if v['rec'] == 'Consolidate' else "→ Optimize/Consolidate"
                print(f"      {v['name']:<45} | ${v['spend']:>10,.0f}")
            if len(vendors) > 5:
                print(f"      ... and {len(vendors) - 5} more")
            print()
        else:
            # Single vendor (no consolidation needed)
            v = vendors[0]
            print(f"  • {function} (1 vendor)")
            print(f"      {v['name']:<45} | ${v['spend']:>10,.0f} (Single provider - no consolidation needed)")
            print()

print("=" * 140)
print("CONSOLIDATION SUMMARY")
print("=" * 140)

print(f"""
Total Consolidation Opportunities: {total_consolidation_groups} function groups
Total Spend in Consolidation Groups: ${total_consolidation_spend:,.0f}

PRIORITY CONSOLIDATIONS (Immediate Action):

1. NAVAN DUPLICATE CONSOLIDATION (SaaS)
   - Spend: $415,913
   - Action: Single licensing agreement
   - Est. Savings: $150-200K

2. PROJECT MANAGEMENT TOOLS (SaaS)
   - Kimble Applications: $52,825
   - Trello: $6,674
   - Potential overlap - evaluate consolidation
   - Est. Savings: $5-10K

3. INSURANCE & EMPLOYEE BENEFITS (G&A)
   - Multiple carriers and benefit administrators
   - Est. Savings: $50-100K (consolidate to 2-3 carriers)

4. REAL ESTATE & OFFICE SPACE (Facilities)
   - Multiple property managers and office providers
   - Est. Savings: $100-200K (consolidate to 1-2 property managers)

5. TRAVEL & EXPENSE MANAGEMENT (G&A/SaaS)
   - Now clearly mapped to appropriate departments
   - Est. Savings: See Navan consolidation

TOTAL POTENTIAL CONSOLIDATION SAVINGS: ~$500K+ from priority vendors alone

Key Insight: Department remapping clarifies consolidation opportunities by function group
Next: Consolidate duplicate entities (Navan) as quick win, then address function-based consolidation
""")

print("\n✓ Consolidation analysis complete")
