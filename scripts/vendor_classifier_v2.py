#!/usr/bin/env python3
"""
Vendor Classification Script v2
Reclassifies all vendors using Config tab department names ONLY
"""

import openpyxl

# Load the output file
file_path = "output/output_file.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb['Vendor Analysis Assessment']

# VALID departments from Config tab
CONFIG_DEPARTMENTS = [
    'Engineering', 'Facilities', 'G&A', 'Legal', 'M&A',
    'Marketing', 'SaaS', 'Product', 'Professional Services',
    'Sales', 'Support', 'Finance'
]

# Mapping rules: vendor name keywords → Config department
department_keywords = {
    'Sales': {
        'keywords': ['salesforce', 'crm', 'hubspot', 'cognism', 'revenue'],
    },
    'Finance': {
        'keywords': ['bank', 'accounting', 'audit', 'tax', 'finance', 'bdo', 'deloitte', 'pwc', 'kpmg', 'ernst', 'grant', 'rsm', 'cpa', 'planful', 'sage', 'ss&c', 'intralinks'],
    },
    'SaaS': {
        'keywords': ['aws', 'amazon web', 'azure', 'google cloud', 'cloud', 'infrastructure', 'hosting', 'network', 'data center', 'software', 'platform', 'saas', 'development', 'tech', 'infosys', 'cloudcrossing', 'kimble', 'telefonica'],
    },
    'Engineering': {
        'keywords': ['engineering', 'developer', 'development'],
    },
    'Product': {
        'keywords': ['monday', 'project management', 'product'],
    },
    'Marketing': {
        'keywords': ['marketing', 'advertising', 'uberflip', 'linkedin', 'mightyhive', 'brand'],
    },
    'Legal': {
        'keywords': ['law', 'legal', 'attorney', 'counsel', 'solicitor'],
    },
    'Professional Services': {
        'keywords': ['recruitment', 'recruiting', 'ats', 'hr', 'talent', 'headhunt', 'staffing', 'consulting', 'advisory', 'strategy', 'advisors', 'houlihan', 'vector'],
    },
    'Facilities': {
        'keywords': ['properties', 'real estate', 'office', 'wework', 'space leasing', 'properties limited'],
    },
    'G&A': {
        'keywords': ['space', 'facilities', 'insurance', 'brokers', 'travel', 'expense', 'navan', 'trip', 'bupa', 'aetna'],
    },
    'M&A': {
        'keywords': ['m&a', 'capital', 'investment', 'advisory services'],
    },
    'Support': {
        'keywords': ['support', 'customer service', 'ticketing'],
    },
}

# Specific vendor overrides for known companies
specific_mappings = {
    'Salesforce Uk Ltd-Uk': ('Sales', 'CRM and sales pipeline management platform'),
    'Navan (Tripactions Inc)': ('G&A', 'Expense management and travel software'),
    'Navan, Inc': ('G&A', 'Expense management and travel software'),
    'Bdo Llp': ('Finance', 'Audit and financial advisory services'),
    'Tog Uk Properties Limited': ('Facilities', 'Office real estate and facilities management'),
    'Cloudcrossing Bvba': ('SaaS', 'Cloud infrastructure and technology services'),
    'Zagrebtower D.O.O.': ('Facilities', 'Office space and facilities management'),
    'Innovent Spaces Private Limited': ('Facilities', 'Office space leasing and design services'),
    'Weking D.O.O.': ('Facilities', 'Office space and business services'),
    'Jensten Insurance Brokers': ('G&A', 'Insurance brokerage and coverage services'),
    'Gpt Space & Co': ('Facilities', 'Office space and facilities management'),
    'Aetna Life And Casualty Ltd': ('G&A', 'Insurance coverage and health benefits'),
    'Rsm Uk Corporate Finance Llp': ('Finance', 'Corporate finance and accounting advisory'),
    'Amazon Web Services Llc': ('SaaS', 'Cloud infrastructure hosting (IaaS)'),
    'Telefonica Global Services Gmbh': ('SaaS', 'Telecommunications and IT services'),
    'Hr Solution International Gmbh': ('Professional Services', 'HR solutions and workforce management'),
    '4I Advisory Services': ('Professional Services', 'Management consulting and advisory services'),
    'Bisley Law Ltd': ('Legal', 'Legal counsel and corporate law services'),
    'Infosys': ('SaaS', 'IT services and software development'),
    'Big Frontier Pty Ltd (Cult Of Monday)': ('Product', 'Project management and work collaboration platform'),
    'Harmonic Group Limited': ('Professional Services', 'Management consulting services'),
    'Wework Singapore Pte. Ltd.': ('Facilities', 'Flexible office space and coworking'),
    'Cloud Technology Solutions Ltd': ('SaaS', 'Cloud infrastructure and technology solutions'),
    'Tmforum': ('SaaS', 'Telecommunications software and standards'),
    'Linkedin Ireland Limited': ('Marketing', 'Social media and professional networking platform'),
    'Kimble Applications Ltd': ('SaaS', 'Project and resource management software'),
    'Sage Uk Limited': ('Finance', 'Accounting and ERP software'),
    'Grant Thornton': ('Finance', 'Audit and financial consulting'),
    'Ss&C Intralinks Inc': ('Finance', 'Financial software and data services'),
    'Veniture D.O.O.': ('G&A', 'Office and business services'),
    'Accutrainee Limited': ('Professional Services', 'Training and talent development services'),
    'Mason Frank International Ltd': ('Professional Services', 'Recruitment and staffing services'),
    'Houlihan Lokey Advisors, Llc': ('Professional Services', 'M&A and financial advisory services'),
    'Vector Capital Management Lp': ('Professional Services', 'Investment and management consulting'),
    'Hubspot Ireland Limited': ('Sales', 'Marketing automation and CRM platform'),
    'Nefron - Obrt Za Poslovne Usluge': ('G&A', 'Business services and operations support'),
    'Planful, Inc.': ('Finance', 'Financial planning and analysis software'),
    'Cognism Limited': ('Sales', 'Sales intelligence and prospecting platform'),
    'Uberflip': ('Marketing', 'Content marketing and engagement platform'),
    'Agram Life Osiguranje D.O.O.': ('G&A', 'Insurance services'),
    'Google Ireland Limited': ('SaaS', 'Cloud services and data analytics platform'),
    'Zuric I Partneri Odvjetnicko Drustvo D.O.O.': ('Legal', 'Legal counsel and corporate law services'),
    'Care Health Insurance Company Limited': ('G&A', 'Health insurance provider'),
    'New Star Networks(Nsn)': ('SaaS', 'IT infrastructure and networking services'),
    'Bupa- Supplier': ('G&A', 'Health insurance and benefits services'),
    'Shree Info System Solutions Pvt Ltd': ('SaaS', 'IT services and software solutions'),
    'Technet It Recruitment': ('Professional Services', 'IT staffing and recruitment services'),
    'Mightyhive Ltd': ('Marketing', 'Digital marketing and advertising services'),
    'Cedar Recruitment Ltd': ('Professional Services', 'Executive recruitment and staffing'),
}

def classify_vendor(vendor_name):
    """Classify vendor using Config departments only"""

    # Check specific mappings first
    if vendor_name in specific_mappings:
        return specific_mappings[vendor_name]

    vendor_lower = vendor_name.lower()

    # Check keyword matches in order of priority
    for dept, rules in department_keywords.items():
        for keyword in rules['keywords']:
            if keyword in vendor_lower:
                # Generate generic description
                desc = f"{dept} vendor and services"
                return (dept, desc)

    # Default fallback
    return ('G&A', 'Business services and operations support')

# Process all vendors
print("=" * 100)
print("VENDOR RECLASSIFICATION WITH CONFIG DEPARTMENTS")
print("=" * 100)

classified_count = 0
department_counts = {}

for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    if vendor_name:
        dept, desc = classify_vendor(vendor_name)

        # Verify department is in Config
        if dept not in CONFIG_DEPARTMENTS:
            print(f"ERROR: Department '{dept}' not in Config for vendor '{vendor_name}'")
            dept = 'G&A'  # Fallback

        ws.cell(row, 2).value = dept
        ws.cell(row, 4).value = desc
        classified_count += 1

        department_counts[dept] = department_counts.get(dept, 0) + 1

        if classified_count % 50 == 0:
            print(f"  Processed {classified_count} vendors...")

print(f"\n✓ Total vendors reclassified: {classified_count}")
print("\nDepartment Distribution:")
total_by_dept = {}
for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    spend = ws.cell(row, 3).value or 0

    if vendor_name and dept:
        if dept not in total_by_dept:
            total_by_dept[dept] = 0
        total_by_dept[dept] += spend

for dept in sorted(CONFIG_DEPARTMENTS):
    count = department_counts.get(dept, 0)
    spend = total_by_dept.get(dept, 0)
    pct = (count / classified_count * 100) if classified_count > 0 else 0
    spend_pct = (spend / sum(total_by_dept.values()) * 100) if sum(total_by_dept.values()) > 0 else 0
    if count > 0:
        print(f"  {dept:20s}: {count:3d} vendors ({pct:5.1f}%) | ${spend:>13,.2f} ({spend_pct:5.1f}%)")

print("\nSample Classifications (first 15):")
print(f"{'Vendor Name':<45} | {'Department':<20} | {'Description':<40}")
print("-" * 110)
for row in range(2, 17):
    vendor = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    desc = ws.cell(row, 4).value
    if vendor:
        print(f"{vendor:<45} | {dept:<20} | {desc:<40}")

# Save
wb.save(file_path)
print(f"\n✓ File saved: {file_path}")
