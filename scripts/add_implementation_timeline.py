#!/usr/bin/env python3
"""
Add Implementation Timeline column to vendor analysis
Estimates realistic timelines in days for executing each recommendation
"""

import openpyxl
import shutil
import os
import time

file_path = "output/output_file.xlsx"
temp_path = "output/output_file_timeline.xlsx"

# Create temp copy
shutil.copy(file_path, temp_path)

wb = openpyxl.load_workbook(temp_path)
ws = wb['Vendor Analysis Assessment']

print("=" * 140)
print("ADDING IMPLEMENTATION TIMELINE COLUMN")
print("=" * 140)

# Add header for timeline column (Column J)
ws.cell(1, 10).value = "Implementation Timeline (Days)"

print("\n✓ Header added: Implementation Timeline (Days)")

# Define timeline logic based on recommendation type
timeline_stats = {
    'Terminate': {'count': 0, 'total_days': 0},
    'Consolidate': {'count': 0, 'total_days': 0},
    'Optimize': {'count': 0, 'total_days': 0}
}

def get_timeline(rec, desc, spend, vendor_name):
    """Calculate implementation timeline based on recommendation and context"""

    desc_lower = desc.lower() if desc else ""
    vendor_lower = vendor_name.lower() if vendor_name else ""

    if rec == 'Terminate':
        # Termination timelines based on complexity
        if spend < 500:
            return 14  # Quick cleanup for very low-spend items
        elif any(kw in desc_lower for kw in ['gym', 'restaurant', 'cafe', 'retail', 'hotel', 'parking']):
            return 30  # Non-core services - standard notice period
        elif any(kw in desc_lower for kw in ['demo', 'trial', 'test']):
            return 7   # Test/demo vendors - immediate termination
        else:
            return 21  # Default termination period

    elif rec == 'Consolidate':
        # Consolidation timelines based on complexity and type
        if 'navan' in vendor_lower:
            return 45  # Navan duplicate - license consolidation, platform migration
        elif any(kw in desc_lower for kw in ['real estate', 'office', 'property', 'workspace']):
            return 120  # Real estate - longest timeline, lease coordination needed
        elif any(kw in desc_lower for kw in ['insurance', 'benefit', 'coverage']):
            return 90   # Insurance - enrollment periods, carrier coordination
        elif any(kw in desc_lower for kw in ['cloud', 'infrastructure', 'hosting', 'aws', 'azure']):
            return 75   # Cloud infrastructure - migration planning and execution
        elif any(kw in desc_lower for kw in ['project', 'task', 'tracking']):
            return 45   # Project management - team migration and training
        elif any(kw in desc_lower for kw in ['consulting', 'advisory']):
            return 60   # Consulting - transition planning with service providers
        elif any(kw in desc_lower for kw in ['recruit', 'staffing', 'ats']):
            return 60   # Recruitment - vendor transition period
        else:
            return 45   # Default consolidation period

    elif rec == 'Optimize':
        # Optimization timelines based on spend level
        if spend > 100000:
            return 45   # High-spend - detailed negotiation, contract review
        elif spend > 10000:
            return 30   # Mid-spend - standard optimization review
        else:
            return 21   # Low-spend - quick optimization check

    return 30  # Default fallback

# Apply timeline to each vendor
timeline_count = 0
for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    rec = ws.cell(row, 5).value or ""
    desc = ws.cell(row, 4).value or ""
    spend = ws.cell(row, 3).value or 0

    if not vendor_name:
        continue

    timeline_days = get_timeline(rec, desc, spend, vendor_name)
    ws.cell(row, 10).value = timeline_days

    if rec in timeline_stats:
        timeline_stats[rec]['count'] += 1
        timeline_stats[rec]['total_days'] += timeline_days

    timeline_count += 1

# Save updated file
wb.save(temp_path)
wb.close()

print(f"\n✓ Timeline added to {timeline_count} vendors")

# Calculate statistics
print("\nTIMELINE SUMMARY BY RECOMMENDATION TYPE:")
print("-" * 140)

total_vendors = 0
total_avg_days = 0
for rec_type in ['Terminate', 'Consolidate', 'Optimize']:
    stats = timeline_stats[rec_type]
    if stats['count'] > 0:
        avg_days = stats['total_days'] / stats['count']
        total_vendors += stats['count']
        total_avg_days += stats['total_days']
        print(f"  {rec_type:<15} | {stats['count']:>3} vendors | Avg: {avg_days:>5.0f} days | Total: {stats['total_days']:>6,} days")

print("-" * 140)
if total_vendors > 0:
    overall_avg = total_avg_days / total_vendors
    print(f"  {'TOTAL':<15} | {total_vendors:>3} vendors | Avg: {overall_avg:>5.0f} days | Total: {total_avg_days:>6,} days")

# Estimate rollout phases
print("\nIMPLEMENTATION ROADMAP:")
print("-" * 140)

terminate_avg = timeline_stats['Terminate']['total_days'] / max(timeline_stats['Terminate']['count'], 1)
consolidate_avg = timeline_stats['Consolidate']['total_days'] / max(timeline_stats['Consolidate']['count'], 1)
optimize_avg = timeline_stats['Optimize']['total_days'] / max(timeline_stats['Optimize']['count'], 1)

print(f"  Phase 1 (Terminate): {int(terminate_avg)} days average")
print(f"    - Quick wins: ~14-30 days (non-core services, low-spend items)")
print(f"    - Can start immediately, parallelizable")
print(f"\n  Phase 2 (Consolidate): {int(consolidate_avg)} days average")
print(f"    - Critical path: ~120 days (real estate consolidation)")
print(f"    - Run in parallel with Phase 1 where possible")
print(f"\n  Phase 3 (Optimize): {int(optimize_avg)} days average")
print(f"    - Negotiation cycles: 21-45 days")
print(f"    - Can run parallel to Phases 1-2")
print(f"\n  Total Critical Path: ~{int(max(terminate_avg, consolidate_avg, optimize_avg))} days (6+ months)")
print(f"  With parallel execution: ~120 days (4 months) for full implementation")

print("\n" + "=" * 140)

# Replace original file with binary write
time.sleep(0.5)
try:
    with open(temp_path, 'rb') as f:
        data = f.read()
    with open(file_path, 'wb') as f:
        f.write(data)
    os.remove(temp_path)
    print(f"✓ Implementation timeline column successfully added to {file_path}")
except PermissionError:
    print(f"⚠ Could not replace original file (locked)")
    print(f"Updated file saved to: {temp_path}")

print("✓ Complete\n")
