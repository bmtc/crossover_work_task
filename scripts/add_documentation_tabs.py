#!/usr/bin/env python3
"""
Add Methodology and Quality Checks documentation tabs to the output spreadsheet.
Makes documentation easily accessible within the Excel file.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import shutil
import os
import time

file_path = "output/output_file.xlsx"
temp_path = "output/output_file_with_docs.xlsx"

# Create temp copy
shutil.copy(file_path, temp_path)

wb = openpyxl.load_workbook(temp_path)

# Remove documentation sheets if they exist
for sheet_name in ['Methodology', 'Quality Checks']:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

print("=" * 140)
print("ADDING DOCUMENTATION TABS TO SPREADSHEET")
print("=" * 140)

# Read documentation files
def read_doc_file(filename):
    with open(filename, 'r', encoding='utf-8') as f:
        return f.read()

methodology_content = read_doc_file('METHODOLOGY_TAB_CONTENT.txt')
quality_content = read_doc_file('QUALITY_CHECKS_TAB_CONTENT.txt')

# Add text to sheet
def add_text_to_sheet(ws, content, title):
    """Add formatted text content to worksheet"""
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")

    ws['A1'] = title
    ws['A1'].fill = title_fill
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(wrap_text=True, vertical='center')

    row = 3
    lines = content.split('\n')

    for line in lines:
        ws[f'A{row}'] = line
        row += 1

    ws.column_dimensions['A'].width = 120

    for row_cell in ws.iter_rows(min_row=3, max_row=row, min_col=1, max_col=1):
        for cell in row_cell:
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

    return row

# Create tabs
print("\n✓ Creating Methodology tab...")
ws_methodology = wb.create_sheet('Methodology')
add_text_to_sheet(ws_methodology, methodology_content, "VENDOR SPEND STRATEGY ASSESSMENT - METHODOLOGY")

print("✓ Creating Quality Checks tab...")
ws_quality = wb.create_sheet('Quality Checks')
add_text_to_sheet(ws_quality, quality_content, "VENDOR SPEND STRATEGY ASSESSMENT - QUALITY CHECKS")

# Save
wb.save(temp_path)
wb.close()

print("\n" + "=" * 140)
print("UPDATING FILE")
print("=" * 140)

# Replace with binary write
time.sleep(0.5)
try:
    with open(temp_path, 'rb') as f:
        data = f.read()
    with open(file_path, 'wb') as f:
        f.write(data)
    os.remove(temp_path)
    print(f"\n✓ Documentation tabs successfully added")
    print(f"\nNew tabs created:")
    print(f"  1. Methodology - Comprehensive methodology documentation")
    print(f"  2. Quality Checks - Executive audit summary with validation results")
except PermissionError:
    print(f"\n⚠ Could not replace original file (locked)")
    print(f"Updated file saved to: {temp_path}")

print("\n✓ Complete")
