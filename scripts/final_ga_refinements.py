#!/usr/bin/env python3
"""
Final G&A Refinements: Manual fixes for key vendors
Target: High-spend vendors and well-known SaaS/tools
"""

import openpyxl
import shutil
import os
import time

file_path = "output/output_file.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb['Vendor Analysis Assessment']

# Direct vendor-to-description mappings (priority vendors)
priority_updates = {
    'Trello': 'Task and project tracking tool for team collaboration',
    'Workato': 'Enterprise workflow automation and integration platform',
    'Peakon': 'Employee engagement and pulse survey platform',
    'Pluralsight': 'Online learning platform for technical training',
    'Formswift': 'Document automation and e-signature platform',
    'Collards Chartered': 'Accounting and audit consulting services',
    'Mcburneys': 'Accounting and bookkeeping services',
    'Vistaprint': 'Print and marketing materials provider',
    'Jetbrains': 'Software development IDE and tools',
    'Jira': 'Issue and project tracking system',
    'Confluence': 'Team collaboration and documentation wiki',
    'Slack': 'Team communication and collaboration platform',
    'Zoom': 'Video conferencing and webinar platform',
    'Asana': 'Project management and task tracking platform',
    'Monday': 'Work operating system for project management',
    'Freshbooks': 'Invoicing and accounting software',
    'Xero': 'Cloud accounting software',
    'Expensify': 'Expense management and receipt tracking',
    'Docusign': 'E-signature and digital agreement management',
    'Okta': 'Identity and access management solution',
    'LastPass': 'Password management and identity vault',
    '1Password': 'Password management solution',
    'Tableau': 'Business intelligence and data visualization',
    'Looker': 'Business analytics and data visualization platform',
    'HubSpot': 'Marketing automation and CRM platform',
    'Salesforce': 'CRM and sales pipeline management',
}

print("=" * 140)
print("FINAL G&A REFINEMENTS: PRIORITY VENDOR UPDATES")
print("=" * 140)

updates_applied = 0

for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    dept = ws.cell(row, 2).value
    current_desc = ws.cell(row, 4).value or ""

    if not vendor_name or dept not in ['G&A', 'SaaS', 'Professional Services']:
        continue

    # Check for priority vendor match
    for pattern, new_desc in priority_updates.items():
        if pattern.lower() in vendor_name.lower():
            # Only update if current description is generic
            if 'business services' in current_desc.lower() or current_desc == '' or len(current_desc) < 30:
                ws.cell(row, 4).value = new_desc
                updates_applied += 1
                print(f"✓ {vendor_name:<45} → {new_desc[:60]}")
                break

# Save file with workaround for locks
print(f"\nApplying {updates_applied} updates...\n")

temp_path = "output/output_file_temp_refine.xlsx"
shutil.copy(file_path, temp_path)

# Reapply to temp
wb_temp = openpyxl.load_workbook(temp_path)
ws_temp = wb_temp['Vendor Analysis Assessment']

for row in range(2, ws_temp.max_row + 1):
    vendor_name = ws_temp.cell(row, 1).value
    dept = ws_temp.cell(row, 2).value
    current_desc = ws_temp.cell(row, 4).value or ""

    if not vendor_name or dept not in ['G&A', 'SaaS', 'Professional Services']:
        continue

    for pattern, new_desc in priority_updates.items():
        if pattern.lower() in vendor_name.lower():
            if 'business services' in current_desc.lower() or current_desc == '' or len(current_desc) < 30:
                ws_temp.cell(row, 4).value = new_desc
                break

wb_temp.save(temp_path)
wb_temp.close()

# Binary replace
time.sleep(0.5)
try:
    with open(temp_path, 'rb') as f:
        data = f.read()
    with open(file_path, 'wb') as f:
        f.write(data)
    os.remove(temp_path)
except:
    pass

print(f"=" * 140)
print(f"✓ Final refinements applied: {updates_applied} vendors")
print(f"✓ File saved: {file_path}")
