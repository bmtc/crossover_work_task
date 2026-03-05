#!/usr/bin/env python3
"""
Phase 2: Improve Vendor Descriptions
Targets vendors with vague descriptions and enhances them with better inference.
"""

import openpyxl

file_path = "output/output_file.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb['Vendor Analysis Assessment']

# Enhanced keyword mappings for better inference
enhanced_mappings = {
    'Jetbrains': ('SaaS', 'Software IDE and development tools'),
    'Formswift': ('SaaS', 'Document automation and e-signature platform'),
    'Vistaprint': ('Marketing', 'Print and marketing materials provider'),
    'Hilton Garden Inn': ('Facilities', 'Hotel and corporate lodging services'),
    'Bureau Veritas': ('Professional Services', 'Quality assurance and testing services'),
    'Sodexo': ('G&A', 'Food service and facilities management'),
    'Jones Lang Lasalle': ('Facilities', 'Real estate advisory and property management'),
    'Gym4You': ('G&A', 'Fitness and wellness facility management'),
    'Calm Achiever': ('Professional Services', 'Employee wellness and training services'),
    'Lane IP': ('Legal', 'Intellectual property and legal services'),
    'Brodomerkur': ('Facilities', 'Shipping and port management services'),
    'Hotel Laguna': ('Facilities', 'Hotel and hospitality services'),
    'Orionw': ('Professional Services', 'Business process outsourcing'),
    'Grad Zagreb': ('G&A', 'Government and municipal administrative services'),
    'Sveu': ('G&A', 'University and educational facility management'),
    'Nefron': ('Professional Services', 'Business consulting and advisory'),
    'Veniture': ('G&A', 'Office space and coworking management'),
    'Agram Life': ('G&A', 'Insurance and health coverage services'),
    'Care Health': ('G&A', 'Health insurance provider'),
    'Harissa': ('G&A', 'Hospitality and food service'),
    'Chamiers Recreation': ('G&A', 'Recreation and membership facility'),
    'Icici Lombard': ('G&A', 'Insurance and risk management'),
    'Golubica Parking': ('G&A', 'Parking and transportation services'),
    'Maniax Melbourne': ('G&A', 'Entertainment and hospitality venue'),
    'Allianz Australia': ('G&A', 'Insurance and workers compensation'),
}

# Additional pattern-based improvements
pattern_improvements = {
    'insurance': ('G&A', lambda vendor: f"Insurance {vendor.split()[0]} and coverage services"),
    'hotel': ('Facilities', lambda vendor: "Hotel and hospitality services"),
    'parking': ('G&A', lambda vendor: "Parking and transportation services"),
    'recreation': ('G&A', lambda vendor: "Recreation and wellness facility"),
    'university': ('G&A', lambda vendor: "University and educational services"),
    'gym': ('G&A', lambda vendor: "Fitness and wellness services"),
    'restaurant': ('G&A', lambda vendor: "Food service and hospitality"),
    'cafe': ('G&A', lambda vendor: "Food service and cafe operations"),
    'dental': ('Professional Services', lambda vendor: "Dental and healthcare services"),
    'medical': ('Professional Services', lambda vendor: "Medical and healthcare services"),
    'clinic': ('Professional Services', lambda vendor: "Medical clinic and healthcare"),
}

print("=" * 100)
print("PHASE 2: VENDOR DESCRIPTION IMPROVEMENT")
print("=" * 100)

improved_count = 0
vendors_to_review = []

for row in range(2, ws.max_row + 1):
    vendor_name = ws.cell(row, 1).value
    current_dept = ws.cell(row, 2).value
    current_desc = ws.cell(row, 4).value

    if not vendor_name:
        continue

    # Check if this vendor needs improvement
    if current_desc and ("business services" in current_desc.lower() or
                         "operations support" in current_desc.lower()):

        # Try exact mapping first
        if vendor_name in enhanced_mappings:
            new_dept, new_desc = enhanced_mappings[vendor_name]
            ws.cell(row, 2).value = new_dept
            ws.cell(row, 4).value = new_desc
            improved_count += 1
            vendors_to_review.append((vendor_name, current_dept, new_dept, new_desc))
            continue

        # Try pattern-based mapping
        vendor_lower = vendor_name.lower()
        improved = False

        for pattern, (suggested_dept, desc_fn) in pattern_improvements.items():
            if pattern in vendor_lower:
                new_desc = desc_fn(vendor_name)
                # Only change dept if suggested dept is specific
                if suggested_dept != 'G&A' or current_dept == 'G&A':
                    ws.cell(row, 2).value = suggested_dept
                    ws.cell(row, 4).value = new_desc
                    improved_count += 1
                    vendors_to_review.append((vendor_name, current_dept, suggested_dept, new_desc))
                    improved = True
                    break

        if improved:
            continue

        # Try substring matching in vendor name for better inference
        vendor_words = vendor_name.lower().split()

        # More sophisticated inference
        if any(word in vendor_lower for word in ['software', 'systems', 'solutions', 'tech', 'app', 'platform']):
            if current_dept == 'G&A':
                ws.cell(row, 2).value = 'SaaS'
                ws.cell(row, 4).value = f"Software platform and technology services"
                improved_count += 1
                vendors_to_review.append((vendor_name, 'G&A', 'SaaS', "Software platform and technology services"))

        elif any(word in vendor_lower for word in ['print', 'design', 'digital', 'marketing', 'advertis']):
            if current_dept == 'G&A':
                ws.cell(row, 2).value = 'Marketing'
                ws.cell(row, 4).value = f"Marketing and {vendor_words[0].lower()} services"
                improved_count += 1
                vendors_to_review.append((vendor_name, 'G&A', 'Marketing', f"Marketing and {vendor_words[0].lower()} services"))

        elif any(word in vendor_lower for word in ['consult', 'advisory', 'training', 'coach', 'development']):
            if current_dept == 'G&A':
                ws.cell(row, 2).value = 'Professional Services'
                ws.cell(row, 4).value = f"Consulting and professional services"
                improved_count += 1
                vendors_to_review.append((vendor_name, 'G&A', 'Professional Services', "Consulting and professional services"))

print(f"\n✓ Improved descriptions: {improved_count} vendors")

if improved_count > 0:
    print(f"\nExamples of improvements made:")
    print(f"{'Vendor Name':<45} | {'From':<20} | {'To':<20} | {'New Description':<40}")
    print("-" * 130)

    for vendor, from_dept, to_dept, new_desc in vendors_to_review[:15]:
        change = f"{from_dept} → {to_dept}" if from_dept != to_dept else from_dept
        print(f"{vendor:<45} | {change:<20} | {new_desc:<40}")

# Save improved file
wb.save(file_path)
print(f"\n✓ File saved with {improved_count} improved descriptions")

# Final statistics
print("\n" + "=" * 100)
print("QUALITY IMPROVEMENT SUMMARY")
print("=" * 100)

vague_remaining = 0
for row in range(2, ws.max_row + 1):
    desc = ws.cell(row, 4).value
    if desc and ("business services" in desc.lower() or "operations support" in desc.lower()):
        vague_remaining += 1

print(f"Vendors improved: {improved_count}")
print(f"Vague descriptions remaining: {vague_remaining} (from 349)")
print(f"Improvement rate: {(improved_count / 349) * 100:.1f}%")
print(f"Description quality increased: 10% → {((386 - vague_remaining) / 386 * 100):.1f}%")
